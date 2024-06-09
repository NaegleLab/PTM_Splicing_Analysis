import pandas as pd
import numpy as np

#analysis libraries
import scipy.stats as stats
from lifelines import KaplanMeierFitter
from Bio import pairwise2

#plotting libraries
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
import matplotlib.cm as cm
import matplotlib.colors as mcl
import networkx as nx
from matplotlib import gridspec, patches
from matplotlib.lines import Line2D
from matplotlib_venn import venn3, venn3_circles
import matplotlib.ticker as mtick
from kstar import plot as kstar_plot

#excel libraries
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#mapper library
from ExonPTMapper import config, mapping

#custom functions
import plot_types
import sys
sys.path.append('../Analysis/')
import stat_utils


class FigureData:
    """
    Loads data needed for figures and used as parameter for other figure classes

    Parameters
    ----------
    min_mods : int, optional
        Minimum number of PTM instances required to consider a modification for plotting. The default is 150.
    axes_label_size : int, optional
        Size of axes labels. The default is 10
    """
    def __init__(self, figshare_dir = './', ptm_data_dir = None, analysis_dir = None):
        self.figshare_dir = figshare_dir
        #where to find projected PTMs
        self.ptm_data_dir = figshare_dir + 'PTM_Projection_Data/' if ptm_data_dir is None else ptm_data_dir
        #where to find information from different databases
        self.database_dir = figshare_dir + 'External_Data/'
        #where analysis data will be saved
        self.analysis_dir = figshare_dir + 'Analysis_For_Paper/' if analysis_dir is None else analysis_dir

        #set global figure parameters
        self.axes_label_size = 10
        self.min_mods = 150

        #load mapper data and add isoform-specific ptm data
        mapper = mapping.PTM_mapper()
        #load isoform-specific ptms
        mapper.isoform_ptms = pd.read_csv(self.ptm_data_dir + '/processed_data_dir/isoform_ptms.csv')
        mapper.calculate_PTMconservation()
        self.mapper = mapper

        #get conversion of PTM subtypes to classes
        self.mod_groups = pd.read_csv(self.database_dir + 'modification_conversion.csv', header = 0)
        

        #separate ptm_info into unique exon/PTM pairs
        exploded_ptms = mapper.explode_PTMinfo()
        exploded_ptms = exploded_ptms.dropna(subset = ['Exon stable ID'])
        self.exploded_ptms = exploded_ptms

        #separate based on modifications (unique PTM/modification class pairs)
        exploded_mods = mapper.ptm_info.copy()
        exploded_mods["Modification Class"] = exploded_mods['Modification Class'].apply(lambda x: x.split(';'))
        exploded_mods = exploded_mods.explode('Modification Class').reset_index()
        exploded_mods = exploded_mods.rename({'index':'PTM'}, axis = 1)
        #exploded_mods = exploded_mods.drop_duplicates(subset = ['Modification', 'PTM'])
        #exploded_mods = exploded_mods.merge(self.mod_groups[['Mod Name', 'Mod Class']], left_on = 'Modification', right_on = 'Mod Name')
        #exploded_mods = exploded_mods.drop(['Mod Name'], axis = 1)
        #exploded_mods = exploded_mods.drop_duplicates()
        self.exploded_mods = exploded_mods

        #set up dictionaries for converting between different nomenclatures
        #converting from PTM identifiers to more readable names
        self.class_name_conversion = {'PHOS':'Phosphorylation', 'UBIQ': 'Ubiquitination', 'ACET':'Acetylation', 'GLCN':'Glycosylation', 'METH':'Methylation',
                 'SUMO':'Sumoylation', 'DIMETH':'Dimethylation', 'NTRY':'Nitrosylation', 'HYDR': 'Hydroxylation', 'TRIMETH':'Trimethylation',
                 'SULF': 'Sulfation', 'PALM':'Palmitoylation', 'GGLU': 'Gamma-carboxyglutamic acid', 'CITR': 'Citrullination',
                 'MYRI':'Myristoylation', 'AMID':'Amidation', 'PYRR':'Pyrrolidone carboxylic acid', 'ADP':'ADP Ribosylation',
                 'PLP':'Pyridoxal phosphate', 'DEAM':'Deamidation', 'GLUT':'Glutathionylation', 'OXOAC': '3-oxoalanine', 
                  'ALLYS':'Allysine', 'TOPA':"2',4',5'-topaquinone", 'CSEA':'Cysteine sulfenic acid', 'PPAN': "O-(pantetheine 4'-phosphoryl)serine",
                 'BIOT':'Biotinylation', 'LIPY':'lipoylation', 'THRX': 'Thyroxine', 'CYSP':'Cysteine persulfide', 'HYPU':'Hypusine',
                 'CSIA':'Cysteine sulfinic acid', 'THIOG': 'Thioglycine', 'DIPH':'Diphthamide', 'DHAS':'2,3-didehydroalanine',
                 'PYRUS':'Pyruvic acid', 'THRN':'Triiodothyronine', 'CETH': 'N6-1-carboxyethyl lysine'}
        self.label_shorthands = {'Phosphorylation':'Ph', 'Ubiquitination':'Ub', 'Acetylation':'Ac', 'Methylation':'Me', 'Glycosylation':'Gly', 'Sumoylation':'Sm', 'Hydroxylation':'Hyd', 'Sulfation':'Sf', 'Nitrosylation':'N', 'Palmitoylation':'Pa', 'Dimethylation': 'DiM', 'Trimethylation':'TriM', 'Crotonylation':'Cr', 'Carboxylation':'Cb', 'Succinylation': 'Sc'}
        #kinase library associated data
        self.kl_kinase_conversion = {'AURKA':'AURA','AURKB':'AURA','AURKC':'AURC','CHEK2':'CHK2','CHEK1':'CHK1','CSNK1A1':'CK1A','CSNK2A1':'CK2A1', 'CSNK2A2':'CK2A2', 'PRKACA':'PKACA',
                    'PRKACB':'PKACB','PRKCA':'PKCA','PRKCB':'PKCB','PRKCE':'PKCE','PRKCI': 'PKCI','PRKG2':'PKG2', 'TRPM7':'CHAK1', 'MAPK14':'P38A',
                    'MAPK1':'ERK2','MAPK3':'ERK1','RPS6KA1':'P90RSK','PRKCQ':'PKCT','PDPK1':'PDK1', 'PRKAA1':'AMPKA1','PRKAA2':'AMPKA2', 'EIF2AK2':'PKR'}

        #annotations from phosphositeplus
        annotations = pd.read_csv(self.database_dir + '/PhosphoSitePlus/Regulatory_sites.gz', sep = '\t',compression = 'gzip', on_bad_lines='skip', header = 2)
        annotations['Substrate'] = annotations['ACC_ID'] + '_' + annotations['MOD_RSD'].apply(lambda x: x.split('-')[0])
        self.annotations = annotations

        #other important variables
        #restrict data to modifications with at least 150 sites
        constitutive_rates = pd.read_csv(self.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)
        constitutive_rates = constitutive_rates[constitutive_rates['Number of Instances in Proteome'] >= 140].sort_values(by = 'Number of Instances in Proteome', ascending = False)
        self.labels = constitutive_rates.index

        #get labels that should be used that are more readable
        #convert modification classes to more readable names for figures
        #labels = []
        #for mod in self.mods_to_keep:
        #    labels.append(self.class_name_conversion[mod])
        #self.labels = labels

    def process_kl_scores(self, kl_scores):
        """
        Process KL scores for plotting
        """
        plt_data_y = kl_scores['Y'].melt(ignore_index = False, var_name = 'kinase', value_name = 'Change in Site Percentile').copy()
        plt_data_y['Mod'] = 'Y'
        plt_data_st = kl_scores['ST'].melt(ignore_index=False, var_name = 'kinase', value_name = 'Change in Site Percentile').copy()
        plt_data_st['Mod'] = 'ST'
        plt_data = pd.concat([plt_data_y, plt_data_st])

        #extract ptm info
        plt_data['PTM'] = plt_data.index.str.split(';')
        plt_data['PTM'] = plt_data['PTM'].apply(lambda x: x[1])

        #add ks data information
         #load psp ks data if not found in data
        if not hasattr(self, 'ks_dataset'):
            ks_dataset = pd.read_csv(self.database_dir + '/PhosphoSitePlus/kinase_substrate_dataset.tsv', sep = '\t')
            #restrict to human interactions
            ks_dataset = ks_dataset[(ks_dataset['KIN_ORGANISM'] == 'human') & (ks_dataset['SUB_ORGANISM'] == 'human')].copy()
            ks_dataset['Source of PTM'] = ks_dataset['SUB_ACC_ID']+'_'+ks_dataset['SUB_MOD_RSD']
            self.ks_dataset = ks_dataset

        #add ks data information
        grouped_ks = self.ks_dataset.copy()
        grouped_ks['PSP Kinase for KL'] = grouped_ks['GENE'].apply(lambda x: self.kl_kinase_conversion[x] if x in self.kl_kinase_conversion else x)
        grouped_ks = grouped_ks[['PSP Kinase for KL', 'Source of PTM']].groupby('Source of PTM')['PSP Kinase for KL'].apply(lambda x: ';'.join([i for i in x if i == i]))
        plt_data = plt_data.merge(grouped_ks, left_on = 'PTM', right_index = True, how = 'left')

        plt_data['Known Interaction'] = plt_data.apply(lambda x: x['kinase'] in x['PSP Kinase for KL'].split(';') if x['PSP Kinase for KL'] == x["PSP Kinase for KL"] else False, axis = 1)
        return plt_data


    def print_summary_information(self):
        print(f"There have been {self.mapper.ptm_info.shape[0]} experimentally observed post-translational modifications across {self.mapper.ptm_info['Gene name'].nunique()} proteins")

        exploded_mods = self.mapper.ptm_info.copy()
        exploded_mods["Modification"] = exploded_mods["Modification"].str.split(';')
        exploded_mods = exploded_mods.explode("Modification")
        print(f"There are {exploded_mods['Modification'].nunique()} unique modification types in the dataset.")
        del exploded_mods


        #fraction of alternative isoforms that do not have PTMs
        tmp = self.mapper.proteins[self.mapper.proteins['UniProt Isoform Type'] == 'Alternative']
        frac_alt_with_no_ptms = tmp[(tmp['Number of PTMs'] == 0) | (tmp['Number of PTMs'].isna())].shape[0]/tmp.shape[0]
        print('Fraction of alternative isoforms with no annotated PTMs: ' + str(frac_alt_with_no_ptms))
        del tmp

        #fraction of ptms that are mapped to alternative isoforms that are not annotated in databases
        ptms_mapped_to_alternative = self.mapper.isoform_ptms[(self.mapper.isoform_ptms['Isoform Type'] == 'Alternative') & (self.mapper.isoform_ptms['Mapping Result'] == 'Success')].copy()
        ptms_mapped_to_alternative['Isoform Sources'] = ptms_mapped_to_alternative['Source of PTM'].apply(lambda x: [ptm.split('_')[0] for ptm in x.split(';')])
        ptms_mapped_to_alternative['PTM from Alternative'] = ptms_mapped_to_alternative.apply(lambda x: x['Isoform ID'] in x['Isoform Sources'] if x['Isoform ID'] == x['Isoform ID'] else False, axis = 1)
        new_ptms_mapped_to_alternative = ptms_mapped_to_alternative[~ptms_mapped_to_alternative["PTM from Alternative"]].shape[0]/ptms_mapped_to_alternative.shape[0]
        print('Fraction of PTMs captured by mapping process that are not found associated with UniProt isoform: ', new_ptms_mapped_to_alternative)

        #overall constitutive rate
        cons_rate = self.mapper.ptm_info[self.mapper.ptm_info['PTM Conservation Score'] == 1].shape[0]/self.mapper.ptm_info.shape[0]
        print(f"The overall constitutive rate is {round(cons_rate*100,2)}%")

        #constitutive rate range
        rates = pd.read_csv(self.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)
        rates = rates[rates['Number of Instances in Proteome'] > 150]
        min_rate = rates['Rate'].min()
        min_mod = rates[rates['Rate'] == min_rate].index[0]
        max_rate = rates['Rate'].max()
        max_mod = rates[rates['Rate'] == max_rate].index[0]
        print(f"The modification with the lowest rate is {min_mod} with a rate of {round(min_rate*100,2)}%")
        print(f"The modification with the highest rate is {max_mod} with a rate of {round(max_rate*100, 2)}%")


        #glycosylation rates
        glycosylation_rate = rates.loc['Glycosylation', 'Rate']
        subtype_rates = pd.read_csv(self.analysis_dir + '/Constitutive_Rates/ByModificationSubtype.csv', index_col = 0)
        oglyc_rate = subtype_rates.loc['O-Glycosylation', 'Rate']
        nglyc_rate = subtype_rates.loc['N-Glycosylation', 'Rate']
        print(f"The overall glycosylation rate is {round(glycosylation_rate*100,2)}%, with O-glycosylation at {round(oglyc_rate*100,2)}% and N-glycosylation at {round(nglyc_rate*100,2)}%")

        #mxe events
        sevents = pd.read_csv(self.ptm_data_dir + '/processed_data_dir/splice_events.csv')
        mxe = sevents[sevents['Event Type'] == 'Mutually Exclusive']
        unique_mxe = mxe.drop_duplicates(subset = ['Canonical Transcript', 'Alternative Transcript', 'Exon ID (Canonical)', 'Exon ID (Alternative)'])
        print(f"We identified {mxe.shape[0]} mutually exclusive exon events, with {unique_mxe.shape[0]} unique events")

        mxe_ptms = self.mapper.alternative_ptms[self.mapper.alternative_ptms['Event Type'] == 'Mutually Exclusive'].copy()
        print(f"In these events, there are {mxe_ptms['Source of PTM'].nunique()} unique PTMs")
        print(f"{round((mxe_ptms[mxe_ptms['Mapping Result'] == 'Success'].shape[0]/mxe_ptms.shape[0])*100, 2)}%' of MXE event PTMs were successfully mapped to isoforms.")

        #altered flanking sequences
        conserved = self.mapper.isoform_ptms[self.mapper.isoform_ptms['Mapping Result'] == 'Success']
        conserved = conserved[conserved['Isoform Type'] == 'Alternative']
        altered_flanks = conserved[conserved['Conserved Flank (Size = 5)'] == 0]
        frac_prospective_with_altered_flank = altered_flanks.shape[0]/conserved.shape[0]
        #get fraction of PTMs that have altered flank but are either found or not found by MS (based on whether tryptic fragment is conserved)
        num_caught_by_MS = altered_flanks[(altered_flanks['Conserved Fragment'] == 0)].shape[0]/altered_flanks.shape[0]
        num_missed_by_MS = altered_flanks[(altered_flanks['Conserved Fragment'] == 1)].shape[0]/altered_flanks.shape[0]
        print(f"Of the {conserved.shape[0]} prospective PTMs, {altered_flanks.shape[0]} ({round(frac_prospective_with_altered_flank*100,2)}%) have altered flanking sequences. Of these, {round(num_caught_by_MS*100,2)}% are found by MS (altered tryptic fragment) and {round(num_missed_by_MS*100,2)}% are not (conserved tryptic fragment).")


        #positions of altered flanking sequences
                #set other variables to none
        altered_flank_positions = pd.read_csv(self.analysis_dir + '/FlankingSequences/PositionOfAlteredFlanks.csv')
        num_with_one_change = altered_flank_positions[altered_flank_positions['Altered_Positions'].apply(lambda x: '1.' in x)].shape[0]
        percent_with_one_change = num_with_one_change/altered_flank_positions.shape[0]
        print(f'{percent_with_one_change*100}% of altered flanking sequences have a change in the +1/-1 position, a total of {num_with_one_change} PTMs')

        
        kl_scores = {}
        kl_scores['Y'] = pd.read_csv(self.analysis_dir + 'FlankingSequences/Kinase_Library/Results/Percentile_Differences_Y.csv', index_col = 0)
        kl_scores['ST'] = pd.read_csv(self.analysis_dir + 'FlankingSequences/Kinase_Library/Results/Percentile_Differences_ST.csv', index_col = 0)
        processed_kl = self.process_kl_scores(kl_scores)
        known = processed_kl[processed_kl['Known Interaction']].shape[0]
        print(f'For the kinase library, analysis we looked at {kl_scores["Y"].shape[0]} events for phosphotyrosine sites and {kl_scores["ST"].shape[0]} events for phosphoserine/threonine sites (total = {kl_scores["Y"].shape[0] + kl_scores["ST"].shape[0]}). \n These include {known} known interactions with phosphorylation site and a kinase in the PhosphoSitePlus database.')
       

class Figure2():
    def __init__(self, data):
        self.data = data
        #get constitutive rates for broad classes
        self.constitutive_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)

        #get splice events impacting each modification
        self.splice_event_fractions = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Splice_Event_Fractions.csv', index_col = 0)

        #load density of ptms in tissue-specific exons
        self.ts_ptm_density = pd.read_csv(self.data.analysis_dir + '/Tissue_Specificity/PTM_Densities.csv', index_col = 0)

        #restrict all data to modifications with at least 150 instances
        self.constitutive_rates = self.constitutive_rates.loc[self.data.labels]
        self.splice_event_fractions = self.splice_event_fractions[self.data.labels]
        self.ts_ptm_density = self.ts_ptm_density.loc[[lab for lab in self.data.labels if lab in self.ts_ptm_density.index]]

        #establish xlimits for all panels
        self.xlim = [-0.5,len(self.data.labels)-0.5]

    def PanelA(self,ax, palette = sns.color_palette('cubehelix', n_colors = 4)):
        """
        Plot the number of instances of each modification type, with bars colored based on the type of molecular being added (small molecule, fatty acid, protein, sugar)
        """
        #color mods based on type
        c = []
        for mod in self.data.labels:
            #small molecule modifications
            if mod in ['Phosphorylation', 'Acetylation', 'Methylation', 'Dimethylation', 'Nitrosylation', 'Trimethylation',
                    'Sulfation', 'Hydroxylation', 'Carboxylation', 'Succinylation','Crotonylation']:
                c.append(palette[0])
            #small protein modifications
            elif mod in ['Ubiquitination', 'Sumoylation']:
                c.append(palette[1])
            #sugar modifications
            elif mod in ['Glycosylation']:
                c.append(palette[2])
            #fatty acid modifications
            elif mod in ['Palmitoylation']:
                c.append(palette[3])

        #construct the legend handles
        legend_elements = [Patch(facecolor=palette[0], edgecolor=palette[0],
                        label='Small Molecule'),
                Patch(facecolor=palette[1], edgecolor=palette[1],
                        label='Protein'),
                Patch(facecolor=palette[2], edgecolor=palette[2],
                        label='Sugar'),
                Patch(facecolor=palette[3], edgecolor=palette[3],
                        label='Fatty Acid')]

        # convert labels to shorthand 
        labels = []
        for mod in self.data.labels:
            labels.append(self.data.label_shorthands[mod])

        #make the barplot (number is log-transformed)
        ax.bar(labels, np.log10(self.constitutive_rates['Number of Instances in Proteome']), color = c)
        ax.set_ylabel('Number\nof PTM\nInstances\n[Log10(N)]', fontsize = self.data.axes_label_size-1)
        ax.set_yticks([0,2,4,6])
        ax.legend(handles = legend_elements, loc = (1.01, 0), title = 'Modification Type', fontsize = self.data.axes_label_size-2, title_fontsize = self.data.axes_label_size-1)
        ax.set_xlim(self.xlim)

    def PanelB(self, ax):
        """
        Modification constitutive rates and their null model comparison plot, which is panel B
        """
        
        #append real constitutive rates
        real_data = self.constitutive_rates['Rate'].reset_index()
        mods_to_plot = self.constitutive_rates.index
        real_data = real_data.rename({'Rate':'Constitutive Rate'}, axis = 1)
        real_data['Residue Type'] = 'Modified (Real)'

        #load null model data
        rate = []
        mod_results = []
        result_type = []
        for mod in mods_to_plot:
            # open file
            with open(self.data.analysis_dir + f'/Null_Model/Null_Constitutive_Rates/{mod}.txt', 'r') as f:

                # write elements of list
                for line in f:
                    rate.append(float(line))
                    mod_results.append(mod)
                    result_type.append('Null (Random)')


            # close the file
            f.close()
            
            
        #construct dataframe with null and real rates
        plt_data = pd.DataFrame({'Modification Class':mod_results, 'Residue Type': result_type, 'Constitutive Rate': rate})

        #convert modification names to readable names (i.e. PHOS to Phosphorylation)
        #mod_class = []
        #for mod in plt_data['Modification']:
        #    mod_class.append(self.data.class_name_conversion[mod])
        #plt_data['Modification Class'] = mod_class



        #append to null data
        plt_data = pd.concat([plt_data, real_data])
        #plt_data = real_data.copy()


        #perform "statistical" test: see fraction of null rates that are less than or greater than real rate, see if it is less than 0.05
        mods = plt_data['Modification Class'].unique()
        comp_result = []
        colors = []
        for mod in mods_to_plot:
            mod_data = plt_data[plt_data['Modification Class'] == mod]
            rand_data = mod_data[mod_data['Residue Type'] == 'Null (Random)']
            mod_rate = mod_data.loc[mod_data['Residue Type'] == 'Modified (Real)', 'Constitutive Rate'].values[0]
            #get fraction of null rates that are more or less than real constitutive rate
            if rand_data.shape[0] == 0:
                comp_result.append(np.nan)
                colors.append(np.nan)
                continue
            else:
                fraction_less = ((mod_rate >= rand_data['Constitutive Rate'])*1).sum()/rand_data.shape[0]
                fraction_more = ((mod_rate <= rand_data['Constitutive Rate'])*1).sum()/rand_data.shape[0]
            #store results
            if fraction_less <= 0.05:
                comp_result.append(fraction_less)
                colors.append('coral')
            elif fraction_more <= 0.05:
                comp_result.append(fraction_more)
                colors.append('cornflowerblue')
            else:
                colors.append(np.nan)
                comp_result.append(np.nan)

        #sort

        #plot bar plot with null rates
        if ax is None:
            fig, ax = plt.subplots(figsize = (6.5,3))
        
        sns.barplot(x = 'Modification Class', y = 'Constitutive Rate', hue = 'Residue Type', data = plt_data, errorbar = 'sd', order = mods_to_plot, ax = ax) 
        ax.legend(loc = (1.01, 0),title = 'Residue Type',fontsize = self.data.axes_label_size-2, title_fontsize = self.data.axes_label_size-1)
        ax.set_yticks([0,1,2,3,4,5])
        if ax is None:
            ax.tick_params(axis = 'x', labelrotation = 35, labelsize = self.data.axes_label_size -1)
        ax.tick_params(axis = 'y', labelsize = self.data.axes_label_size -1)


        #add stat annotations
        for i in range(len(comp_result)):
            comp = comp_result[i]
            if comp == comp:
                x1 = i - 0.25
                x2 = i + 0.25
                y = 0.9
                h = 0.015
                ax.plot([x1, x1, x2, x2], [y, y+h, y+h, y], lw=1, color = colors[i])
                if comp == 0:
                    annot = '**'
                elif comp <= 0.05:
                    annot = '*'
                ax.text((x1+x2)*.5, y+0.02, annot, ha='center', va='center', fontsize = 10, color = colors[i])

        # convert labels to shorthand 
        labels = []
        for mod in self.data.labels:
            labels.append(self.data.label_shorthands[mod])

        ax.set_xticklabels(labels)

        #format plot
        ax.set_ylim([0.4,1])
        ax.set_yticks([0.4,0.6,0.8, 1])
        ax.set_xlabel('')
        ax.set_ylabel('Constitutive\nRate', fontsize = self.data.axes_label_size-1)
        ax.set_xlim(self.xlim)


    def PanelC(self, ax):
        """
        Plot which events are leading to ptm exclusion
        """
        #sevents plot
        plt_data = self.splice_event_fractions.T.copy()
        #combine alternative splice site events into one group
        plt_data['Alternative Splice Site'] = plt_data[["3' ASS", "3' and 5' ASS", "5' ASS"]].sum(axis = 1)
        plt_data = plt_data.drop(["3' ASS", "3' and 5' ASS", "5' ASS"], axis = 1)
        plt_data = plt_data.rename({'Skipped':'Skipped Exon','Mutually Exclusive':'Mutually Exclusive Exon'}, axis = 1)
        plt_data = plt_data.T

        # convert labels to shorthand
        labels = []
        for mod in self.data.labels:
            labels.append(self.data.label_shorthands[mod])

        plot_types.stackedBar(labels, plt_data[self.data.labels].values, ax = ax, colormap = 'colorblind', legend = list(plt_data.index),
                leg_loc = [1.01, 0], legend_fontsize=self.data.axes_label_size-2, title_fontsize = self.data.axes_label_size-1)
        ax.set_ylim([0,1])
        ax.set_ylabel('Splice Event\nFraction', fontsize = self.data.axes_label_size-1)

        # set xlim 
        ax.set_xlim(self.xlim)

    def PanelD(self, ax):
        """
        Tissue specificity plot
        """
        #tissue specificity plot
        ts_plt_data = self.ts_ptm_density.copy()
        for lab in self.data.labels:
            if lab not in ts_plt_data.index:
                ts_plt_data.loc[lab] = np.nan
        #color bars based on whether PTMs are more or less dense than expected
        c = []
        palette = ['cornflowerblue', 'coral']
        for mod in self.data.labels:
            if ts_plt_data.loc[mod, 'Normalized PTM Density'] == ts_plt_data.loc[mod, 'Normalized PTM Density']:
                if ts_plt_data.loc[mod, 'Normalized PTM Density'] > 1:
                    c.append(palette[1])
                else:
                    c.append(palette[0])
                    
        legend_elements = [Patch(facecolor=palette[1], edgecolor=palette[1],
                                label='Enriched (> 1)'),
                        Patch(facecolor=palette[0], edgecolor=palette[0],
                                label='Depleted (< 1)')]
        
        labels = []
        for mod in self.data.labels:
            labels.append(f'{mod} ({self.data.label_shorthands[mod]})')
        
        for lab in self.data.labels:
            if lab not in ts_plt_data.index:
                ts_plt_data.loc[lab] = np.nan

        ax.bar(labels, ts_plt_data.loc[self.data.labels, 'Normalized PTM Density'], color = c)
        ax.set_ylabel('Normalized\nPTM Density in\nTissue-Specific\nExons', fontsize = self.data.axes_label_size-1)
        ax.set_xlim(self.xlim) 
        ax.set_yticks([0,0.5, 1,1.5,2])
        ax.legend(handles = legend_elements, loc = (1.01,0), fontsize = self.data.axes_label_size-2, title_fontsize = self.data.axes_label_size-1)

    def generate_figure(self, panel_label_size = 20, fig_save_dir = None, fig_type = 'svg'):
        """
        Generate complete panels

        """
        # setup figures
        fig, ax = plt.subplots(figsize = (5, 5), nrows = 4, height_ratios = [0.6,1,1,1])
        fig.subplots_adjust(hspace = 0.35)

        panel_label_position = -3.7
        #rate plot
        self.PanelA(ax[0])
        ax[0].text(panel_label_position, 6, 'A', fontsize = panel_label_size, fontweight = 'bold')
        ax[0].tick_params(axis = 'y', labelsize = self.data.axes_label_size -1)
        ax[0].tick_params(axis = 'x', labelsize = self.data.axes_label_size -1)
        self.PanelB(ax[1])
        ax[1].text(panel_label_position, 0.9, 'B', fontsize = panel_label_size, fontweight = 'bold')
        ax[1].tick_params(axis = 'y', labelsize = self.data.axes_label_size -1)
        ax[1].tick_params(axis = 'x', labelsize = self.data.axes_label_size -1)
        self.PanelC(ax[2])
        ax[2].text(panel_label_position, 0.9, 'C', fontsize = panel_label_size, fontweight = 'bold')
        ax[2].tick_params(axis = 'y', labelsize = self.data.axes_label_size -1)
        ax[2].tick_params(axis = 'x', labelsize = self.data.axes_label_size -1)
        self.PanelD(ax[3])
        ax[3].text(panel_label_position, 3, 'D', fontsize = panel_label_size, fontweight = 'bold')
        ax[3].tick_params(axis = 'y', labelsize = self.data.axes_label_size -1)
        ax[3].tick_params(axis = 'x', labelsize = self.data.axes_label_size -1)
        ticks = plt.xticks(rotation = 35, ha = 'right')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'Figure2.{fig_type}', format = fig_type, bbox_inches = 'tight', dpi = 300)




class Figure3:
    def __init__(self, data):
        self.data = data

        self.flank_rates = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/Mod_Specific_Alteration_Rates.csv', index_col = 0)
        self.window5_flanks = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/Window5_FlankingSequences.csv', index_col = 0)
        #get labels
        if self.data.labels is None:
            labels = []
            for mods in self.data.mods_to_keep:
                labels.append(self.data.conversion[mods])
            self.data.labels = labels

        self.sequence_similarity = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/SequenceSimilarity.csv')

        self.kl_scores = {}
        self.kl_scores['Y'] = pd.read_csv(self.data.analysis_dir + 'FlankingSequences/Kinase_Library/Results/Percentile_Differences_Y.csv', index_col = 0)
        self.kl_scores['ST'] = pd.read_csv(self.data.analysis_dir + 'FlankingSequences/Kinase_Library/Results/Percentile_Differences_ST.csv', index_col = 0)
        self.motif_change_title = 'Change in Site Percentile'

        #load psp ks data if not found in data
        if not hasattr(self.data, 'ks_dataset'):
            ks_dataset = pd.read_csv(self.data.database_dir + '/PhosphoSitePlus/kinase_substrate_dataset.tsv', sep = '\t')
            #restrict to human interactions
            ks_dataset = ks_dataset[(ks_dataset['KIN_ORGANISM'] == 'human') & (ks_dataset['SUB_ORGANISM'] == 'human')].copy()
            ks_dataset['Source of PTM'] = ks_dataset['SUB_ACC_ID']+'_'+ks_dataset['SUB_MOD_RSD']
            self.data.ks_dataset = ks_dataset



    def PanelB(self, ax, palette = sns.color_palette('colorblind'), leg_loc = [-0.8, 1.4], legend_fontsize = 8, fig_save_dir = None):
        """
        Plot figure3B, which shows the fraction of PTMs with conserved flanking sequences and tryptic fragments (overall)

        Parameters
        ----------
        ax : matplotlib axis, optional
            Axis to plot on. The default is None.
        palette : list, optional
            Color palette to use. The default is sns.color_palette('colorblind').
        leg_loc : list, optional
            Location of legend. The default is [-0.8, 1.4].

        Returns
        -------
        None.
        """
        conserved = self.data.mapper.isoform_ptms[self.data.mapper.isoform_ptms['Mapping Result'] == 'Success']
        conserved = conserved[conserved['Isoform Type'] == 'Alternative']
        #get fraction of PTMs that have altered flank but are either found or not found by MS (based on whether tryptic fragment is conserved)
        num_caught_by_MS = conserved[(conserved['Conserved Flank (Size = 5)'] == 0) & (conserved['Conserved Fragment'] == 0)].shape[0]/conserved.shape[0]
        num_missed_by_MS = conserved[(conserved['Conserved Flank (Size = 5)'] == 0) & (conserved['Conserved Fragment'] == 1)].shape[0]/conserved.shape[0]
        total_conserved = conserved[conserved['Conserved Flank (Size = 5)'] == 1].shape[0]/conserved.shape[0]

        
        #plot stacked bar with fractions and annotate with number of PTMs
        plot_types.stackedBarH('Test', [[total_conserved], [num_missed_by_MS],[num_caught_by_MS]], colormap = 'colorblind', y_label = 'Fraction of Prospective PTMs', ax = ax, 
                    legend = ['Unaltered Regulatory Region', 'Altered Regulatory Region Only', 'Altered Regulatory Region and Tryptic Fragment'], leg_loc = leg_loc, legend_fontsize = legend_fontsize)
        ax.set_yticks([])

        #add specific numbers and percent to plot
        ax.annotate(f"{conserved[conserved['Conserved Flank (Size = 5)'] == 1].shape[0]:,} ({round(total_conserved*100,2)}%)", (0.4, 0.2), (0.1, 0.5), arrowprops = {'width':0.2,'headwidth':5},
                fontsize = 9, ha = 'center')
        ax.annotate(f"{conserved[(conserved['Conserved Flank (Size = 5)'] == 0) & (conserved['Conserved Fragment'] == 1)].shape[0]:,} ({round(num_missed_by_MS*100,2)}%)", 
                    (1-num_missed_by_MS-num_caught_by_MS, 0.2), (0.7, 0.5), arrowprops = {'width':0.2,'headwidth':5, 'facecolor':palette[1]},fontsize = 9, ha = 'center')
        ax.annotate(f"{conserved[(conserved['Conserved Flank (Size = 5)'] == 0) & (conserved['Conserved Fragment'] == 0)].shape[0]:,} ({round(num_caught_by_MS*100,2)}%)", 
                    (1, 0.2), (1.2, 0.5), arrowprops = {'width':0.2,'headwidth':5, 'facecolor':palette[2]},fontsize = 9, ha = 'center')

        
        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3B.svg', dpi = 300, bbox_inches = 'tight')

    def PanelC(self, ax, as_percent = True, leg_loc = [0.1, 0.45], fig_save_dir = None):
        """
        Make Figure 3 Panel C, which contains the modification alteration rates for each modification and window size

        Parameters
        ----------
        ax : matplotlib axis, optional
            Axis to plot on. The default is None.
        leg_loc : list, optional
            Location of legend. The default is [0.1, 0.45].
        """
        print('Making Figure 3C')
        #get fraction of PTMs with an altered flanking sequence specifically when that flanking sequence is considered (i.e. if window size is 3, subtract number with altered flank at window size 1 and 2)
        shifted_mod_flanks = self.flank_rates.copy()
        for i in list(range(shifted_mod_flanks.shape[1]-1,0, -1)):
            shifted_mod_flanks.iloc[:,i] = shifted_mod_flanks.iloc[:,i] - shifted_mod_flanks.iloc[:,i-1]
        flank_sizes = ['1', '2', '3', '4' , '5']
        plt_data = shifted_mod_flanks.loc[self.data.labels, flank_sizes]

        #make figure if not provided
        if ax is None:
            fig, ax = plt.subplots(figsize = (3,3))

        #plot stacked bar for each modification, with window size colored
        plot_types.stackedBar(self.data.labels, plt_data.T.values, ax = ax, y_label = 'Altered Flank Rate', legend = flank_sizes,
                            leg_loc = leg_loc, legend_title = 'Window Size', colormap = 'colorblind', legend_fontsize=8)
        ax.set_ylim([0,0.1])
        ax.set_yticks([0,0.02,0.04,0.06,0.08,0.1])

        #convert to percent
        plt_data = plt_data * 100
        #plot
        plot_types.stackedBar(self.data.labels, plt_data.T.values, ax = ax, y_label = 'Altered Flank Rate', legend = flank_sizes,
                leg_loc = leg_loc, legend_title = 'Window Size', colormap = 'colorblind', legend_fontsize=8)
        ax.set_ylim([0,10])
        ax.set_yticks([0,2,4,6,8,10])
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(decimals = 0))
        ax.set_ylabel('Percent of PTMs\n with Altered Flank', fontsize = self.data.axes_label_size)

        plt.xticks(rotation = 45, ha = 'right', fontsize = self.data.axes_label_size-2)


        #annotate with the number of ptms on top of the bar
        ticks = ax.get_xticks() 
        plt_data2 = self.window5_flanks.loc[self.data.labels]
        for i, mod in zip(range(plt_data2.shape[0]), plt_data2.index):
            num_altered_mods = plt_data2.loc[mod,'Number of PTMs with Altered Flanking Sequence']
            if as_percent:
                ax.annotate(f'{num_altered_mods}',(ticks[i], plt_data2.loc[mod,'Fraction of PTMs with Altered Flanking Sequence']*100+0.1), rotation = 90, 
                        ha = 'center', va = 'bottom')
            else:
                ax.annotate(f'{num_altered_mods}',(ticks[i], plt_data2.loc[mod,'Fraction of PTMs with Altered Flanking Sequence']+0.001), rotation = 90, 
                        ha = 'center', va = 'bottom')
                
        
        
        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3C.svg', dpi = 300, bbox_inches = 'tight')
                
    def newCauseNames(self, altered_flanks_with_cause):
        """
        Given a dataframe of altered flanks which indicate the cause of alteration in the column 'Cause', rename the column names

        Parameters
        ----------
        altered_flanks_with_cause : pandas dataframe
            dataframe containing information about altered flanks, including the cause of alteration in the column 'Cause'
        
        Returns
        -------
        new_cause : list
            list of new cause names
        """
        new_cause = []
        for i, row in altered_flanks_with_cause.iterrows():
            if 'ASS' in row['Cause']:
                if row['Cause'] == 'ASS (Exon with PTM)':
                    new_cause.append('ASS (Internal)')
                else:
                    new_cause.append('ASS (Adjacent)')
            elif 'MXE' in row['Cause']:
                if row['Cause'] == 'MXE (Exon with PTM)':
                    new_cause.append('MXE (Internal)')
                else:
                    new_cause.append('MXE (Adjacent)')
            elif 'Skipped (' in row['Cause']:
                new_cause.append('Skipped (Adjacent)')
            elif 'Inserted' in row['Cause']:
                new_cause.append('Retained (Adjacent)')
            else:
                new_cause.append(np.nan)
        return new_cause


    def PanelD(self, fig_save_dir = None):
        """
        Make Figure 3 Panel D, which indicates how PTM flanks are altered

        Parameters
        ----------
        gs : matplotlib gridspec, optional
            Gridspec to plot on. The default is None.
        fig : matplotlib figure, optional
            Figure to plot on. Required if gridspec is provided. The default is None.
        """
        print('Plotting Figure 3D')
        altered_flanks = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/cause_of_alteration.csv')
        altered_flanks['Cause'] = altered_flanks['Cause'].apply(lambda x: x.split(','))
        altered_flanks = altered_flanks.explode('Cause')


        isoform_ptms = self.data.mapper.isoform_ptms.copy()
        cols_to_collapse = ['Gene stable ID', 'Alternative Transcript',  'Exon ID (Alternative)', 'Exon ID (Canonical)', 'Second Exon']
        if 'TRIFID Score' in self.data.mapper.isoform_ptms.columns:
            cols_to_collapse.append('TRIFID Score')

        #collapse rows on ptm and isoform id
        group_cols = [col for col in self.data.mapper.isoform_ptms.columns if col not in cols_to_collapse]
        isoform_ptms = isoform_ptms.groupby(group_cols).agg(lambda x: ';'.join(np.unique([str(y) for y in x if y == y]))).reset_index()
        isoform_ptms = isoform_ptms.replace('',np.nan)

        #rename causes and remove any that are not in the list
        altered_flanks['Causes for Plot'] = self.newCauseNames(altered_flanks)
        altered_flanks = altered_flanks.dropna(subset = 'Causes for Plot')

        #separate cause information into event type and location (within the same exon or adjacent exon)
        altered_flanks['Location of Alteration'] = altered_flanks['Causes for Plot'].apply(lambda x: x.split(' ')[1].strip('(').strip(')'))
        altered_flanks['Event Type'] = altered_flanks['Causes for Plot'].apply(lambda x: x.split(' ')[0])
        altered_flanks['Event Type'] = altered_flanks['Event Type'].apply(lambda x: 'Skipped/Retained' if x == 'Skipped' or x == 'Retained' else x)

        fig, ax = plt.subplots(figsize = (3,1.5), nrows = 2, sharex = True)
        fig.subplots_adjust(hspace = 0)


        #plot stacked bar plots with fraction of causes for altered flanks
        plot_types.stackedBarH('Location', [[y] for y in altered_flanks.groupby('Location of Alteration').size()/altered_flanks.shape[0]], ax = ax[0], colormap = 'terrain',legend = ['Adjacent', 'Internal'], leg_loc = 'center', legend_bbox_to_anchor=[0.5, 1.3], legend_type = 'horizontal')
        ax[0].set_xticks([])
        plot_types.stackedBarH('Event Type', [[y] for y in altered_flanks.groupby('Event Type').size()/altered_flanks.shape[0]], figsize = (2,0.25), ax = ax[1],
                            legend = ['Alternative Splice Site', 'Mutually Exclusive Exons', 'Skipped/Retained'], leg_loc = 'center', legend_bbox_to_anchor=[0.5, -1.3], legend_type = 'vertical', y_label = 'Fraction of PTMs')
        

        
        
        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3D.svg', dpi = 300, bbox_inches = 'tight')
            
    def PanelE(self, ax = None, fig_save_dir = None):
        """
        Make Figure 3E: The sequence similarity between flanking sequences that are different between canonical and alternative flanking sequences

        Parameters
        """
        print('Making Figure3E')
        #make boxplots
        if ax is None:
            fig, ax = plt.subplots(figsize = (1.5,1.5))
        
        #construct plot
        sns.boxplot(x= 'Window Size', y = 'Similarity', data = self.sequence_similarity, color = 'lightblue')
        ax.set_xlabel('Window Size')
        ax.set_ylabel('Sequence Identity')
        ax.set_yticks([0,20,40,60,80,100])

        #convert ticks to percentages
        import matplotlib.ticker as mtick
        fmt = '%.0f%%' # Format you want the ticks, e.g. '40%'
        yticks = mtick.FormatStrFormatter(fmt)
        ax.yaxis.set_major_formatter(yticks)

        
        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3E.svg', dpi = 300, bbox_inches = 'tight')

    def process_kl_scores(self):
        """
        Process KL scores for plotting
        """
        plt_data_y = self.kl_scores['Y'].melt(ignore_index = False, var_name = 'kinase', value_name = 'Change in Site Percentile').copy()
        plt_data_y['Mod'] = 'Y'
        plt_data_st = self.kl_scores['ST'].melt(ignore_index=False, var_name = 'kinase', value_name = 'Change in Site Percentile').copy()
        plt_data_st['Mod'] = 'ST'
        plt_data = pd.concat([plt_data_y, plt_data_st])

        #extract ptm info
        plt_data['PTM'] = plt_data.index.str.split(';')
        plt_data['PTM'] = plt_data['PTM'].apply(lambda x: x[1])

        #add ks data information
        grouped_ks = self.data.ks_dataset.copy()
        grouped_ks['PSP Kinase for KL'] = grouped_ks['GENE'].apply(lambda x: self.data.kl_kinase_conversion[x] if x in self.data.kl_kinase_conversion else x)
        grouped_ks = grouped_ks[['PSP Kinase for KL', 'Source of PTM']].groupby('Source of PTM')['PSP Kinase for KL'].apply(lambda x: ';'.join([i for i in x if i == i]))
        plt_data = plt_data.merge(grouped_ks, left_on = 'PTM', right_index = True, how = 'left')

        plt_data['Known Interaction'] = plt_data.apply(lambda x: x['kinase'] in x['PSP Kinase for KL'].split(';') if x['PSP Kinase for KL'] == x["PSP Kinase for KL"] else False, axis = 1)
        self.kl_plt_data = plt_data

    def PanelF(self, ax = None, fig_save_dir = None):
        """
        Make Figure 3F
        """
        if not hasattr(self, 'kl_plt_data'):
            self.process_kl_scores()

        print("Making Figure 3F")

        palette = sns.color_palette('colorblind')
        if ax is None:
            fig, ax = plt.subplots(figsize = (3,2))


        #get number of known and unknown interactions
        known = self.kl_plt_data[self.kl_plt_data['Known Interaction']]
        unknown = self.kl_plt_data[~self.kl_plt_data['Known Interaction']]

        #plot histogram
        hist = ax.hist(unknown['Change in Site Percentile'], alpha = 0.5, density = True, label = f'No (n={unknown.shape[0]:,})', color = palette[3])
        hist = ax.hist(known['Change in Site Percentile'], alpha = 0.5, density = True, label = f'Yes (n={known.shape[0]:,})', color = palette[2])
        ax.legend(title = 'Known Interaction', ncols = 2, loc = 'center', bbox_to_anchor = (0.5, -0.5), fontsize = 10, title_fontsize = 10)

        ax.set_xlabel(self.motif_change_title, fontsize = self.data.axes_label_size)
        ax.set_ylabel('Density', fontsize = self.data.axes_label_size)
        ax.axvline(0, linestyle = 'dashed', color = 'black', alpha = 0.4)
        ax.set_xlim([-70,70])
        ax.text(8,0.16,'Favors Alternative\nSequence', ha = 'left', fontsize = 10)
        ax.text(-8,0.16,'Favors Canonical\nSequence', ha = 'right', fontsize = 10)

        #add number of ptms and kinases
        ax.annotate(f'{self.kl_plt_data["PTM"].nunique()} PTMs', (68,0.1), ha = 'right', fontsize = self.data.axes_label_size -1)
        ax.annotate(f'{self.kl_plt_data["kinase"].nunique()} Kinases', (68,0.085), ha = 'right', fontsize = self.data.axes_label_size - 1)
        ax.annotate(f'{self.kl_plt_data.shape[0]:,} Possible\nInteractions', (68,0.05), ha = 'right', fontsize = self.data.axes_label_size - 1)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3F.svg', dpi = 300, bbox_inches = 'tight')



    def PanelG_Transcript(self, gs = None, fig = None, fig_save_dir = None):
        """
        Making Figure 3G
        """
        print("Making Figure 3G, Transcript Diagram")
        plot_types.compareTranscripts(self.data.mapper, canonical_transcript = 'ENST00000382743', alternative_transcript = 'ENST00000524564', ptm = 'Q9NTG7-1_S159', fontsize = 18, gs= gs, fig = fig)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3G_transcript.svg', dpi = 300, bbox_inches = 'tight')


    def PanelG_Network(self, substrate_node_color = 'darkgray', kinase_node_color = 'white', gs = None, fig = None, fig_save_dir = None):
        """
        Making Figure 3H
        """
        if not hasattr(self, 'kl_plt_data'):
            self.process_kl_scores()

        print("Making Figure 3G, Network")

        sirt3_scores = self.kl_plt_data[self.kl_plt_data['PTM'] == 'Q9NTG7_S159']
        sirt3_scores = sirt3_scores[['kinase', 'Known Interaction', 'Change in Site Percentile']]
        sirt3_scores = sirt3_scores[(sirt3_scores['kinase'].str.contains('CDK')) & 
                    ~(sirt3_scores['kinase'].str.contains('CDKL')) &
                    (sirt3_scores['Change in Site Percentile'].apply(abs) > 8)]
        sirt3_scores['substrate'] = 'SIRT3\nS159'

        #setup subplot
        if gs is None:
            fig, ax = plt.subplots(figsize = (3,3), nrows = 2, height_ratios = [1,0.05])
            fig.subplots_adjust(wspace = 0.02)
        else:
            #construct subplots within gridspec
            gs2 = gridspec.GridSpecFromSubplotSpec(2, 1, subplot_spec=gs, wspace = 0.02, height_ratios = [1.5,0.05])
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1])]

        #create directed graph and add edges
        G = nx.DiGraph()
        for i, row in sirt3_scores.iterrows():
            G.add_edge(row['kinase'],row['substrate'],weight=row['Change in Site Percentile'])

        #color edges based on sign of difference (red for increase, blue for decrease)
        #normalizer = mcl.Normalize(vmin = -3, vmax = 3)
        normalizer = mcl.Normalize(vmin = -20, vmax = 20)
        cmap = cm.ScalarMappable(normalizer,cmap = 'coolwarm')
        # Define the edge colors based on the sign of the weight
        edge_colors = [cmap.to_rgba(w) for u, v, w in G.edges.data('weight')]

        # Define the edge widths based on the absolute value of the weight
        edge_widths = [2 for u, v, w in G.edges.data('weight')]

        node_colors = []
        node_edge_colors = []
        for node in G.nodes:
            if 'SIRT3' in node:
                node_colors.append(substrate_node_color)
                node_edge_colors.append('white')
            elif node == 'CDK1':
                node_colors.append(kinase_node_color)
                node_edge_colors.append('green')
            else:
                node_colors.append(kinase_node_color)
                node_edge_colors.append('black')

        # Draw the graph with matplotlib
        pos = nx.circular_layout(G)
        pos['SIRT3\nS159'] = np.array([0,0])
        nodes = nx.draw_networkx_nodes(G, pos, node_size = 600, node_color = node_colors, edgecolors = node_edge_colors, ax = ax[0])
        nx.draw_networkx_labels(G, pos, font_size = 7, ax = ax[0])
        nx.draw_networkx_edges(G, pos, edge_color=edge_colors, width=edge_widths, ax = ax[0])
        fig.colorbar(cmap,cax=ax[1], orientation='horizontal', label=self.motif_change_title)
        ax[0].axis('off')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'Figure3G_network.svg', dpi = 300, bbox_inches = 'tight')




class SupplementaryFigure2:
    def __init__(self, data):
        #point to figure data class
        self.data = data



        isoform_ptms = data.mapper.proteins.dropna(subset = 'UniProt Isoform Type')
        isoform_ptms = isoform_ptms[['UniProtKB/Swiss-Prot ID', 'UniProt Isoform Type', 'Number of PTMs']]
        isoform_ptms = isoform_ptms.replace(np.nan, 0)
        self.isoform_ptms = isoform_ptms


    def PanelA_table(self, ax = None, panel_save_dir = None):
        #data summary
        num_genes = self.data.mapper.genes.shape[0]
        num_transcripts = self.data.mapper.transcripts.shape[0]
        num_ensembl_isoforms = self.data.mapper.isoforms[(self.data.mapper.isoforms['Isoform Type'] == 'Alternative') | (self.data.mapper.isoforms['Isoform Type'].isna())].shape[0]
        num_uniprot_proteins = self.data.mapper.proteins['UniProtKB/Swiss-Prot ID'].nunique()
        num_uniprot_isoforms = self.data.mapper.proteins[self.data.mapper.proteins['UniProt Isoform Type'] == 'Alternative'].shape[0]
        num_ptms = self.data.mapper.ptm_coordinates.shape[0]

        #proteins with PTMs
        with_ptms = self.data.mapper.proteins[self.data.mapper.proteins['Number of PTMs'] > 0]
        num_proteins_with_ptms = with_ptms['UniProtKB/Swiss-Prot ID'].nunique()
        frac_proteins_with_ptms = num_proteins_with_ptms/num_uniprot_proteins
        isoforms_with_ptms = with_ptms[with_ptms['UniProt Isoform Type'] == 'Alternative']
        num_isoforms_with_ptms = isoforms_with_ptms.shape[0]
        frac_isoforms_with_ptms = num_isoforms_with_ptms/num_uniprot_isoforms

        summary = pd.DataFrame({'Genes': [num_genes,'Ensembl'], 'Transcripts': [num_transcripts, 'Ensembl'],'Alternative Isoforms (Ensembl)':[num_ensembl_isoforms, 'Ensembl'], 'Proteins':[num_uniprot_proteins, 'UniProtKB'], 'Alternative Isoforms (Uniprot)':[num_uniprot_isoforms, 'UniProtKB'], 'Post Translational Modifications (PTMs)': [num_ptms, 'ProteomeScout/PhosphoSitePlus'], 'Genes/Proteins with PTMs':[f'{num_proteins_with_ptms} ({round(frac_proteins_with_ptms*100,1)}%)', 'ProteomeScout/PhosphoSitePlus'], 'UniProtKB Alternative Isoforms with annotated PTMs':[f'{num_isoforms_with_ptms} ({round(frac_isoforms_with_ptms*100,1)}%)', 'ProteomeScout/PhosphoSitePlus']}, index = ['Number of Instances', 'Source'])
        
        if panel_save_dir is not None:
            summary.T.to_csv(panel_save_dir + 'SFig2_data_summary.csv')

        self.summary = summary.T

    def PanelB(self, ax = None, panel_save_dir = None):
        """
        Violin plot
        """
        #construct figure
        if ax is None:
            fig, ax = plt.subplots(figsize = (2.5,3))

        #add post-mapping information
        found_ptms = self.data.mapper.isoform_ptms[self.data.mapper.isoform_ptms['Mapping Result'] == 'Success'].copy()
        found_ptms_canonical = found_ptms[found_ptms['Isoform Type'] == 'Canonical']
        found_ptms_alternative = found_ptms[found_ptms['Isoform Type'] == 'Alternative']
        found_ptms_can_count = found_ptms_canonical.groupby('Isoform ID').size().reset_index()
        found_ptms_can_count['UniProt Isoform Type'] = 'Canonical'
        found_ptms_alt_count = found_ptms_alternative.groupby('Isoform ID').size().reset_index()
        found_ptms_alt_count['UniProt Isoform Type'] = 'Alternative'
        found_ptms_count = pd.concat([found_ptms_can_count, found_ptms_alt_count])
        found_ptms_count['Source'] = 'Post-mapping'
        found_ptms_count = found_ptms_count.rename(columns = {0:'Number of PTMs'})

        self.isoform_ptms['Source'] = 'Pre-mapping'
        self.isoform_ptms = self.isoform_ptms.rename(columns = {'UniProtKB/Swiss-Prot ID':'UniProtKB/Swiss-Prot ID'})
        self.isoform_ptms = pd.concat([self.isoform_ptms, found_ptms_count])
        
        #plot violin plot
        palette = sns.color_palette('colorblind')
        box = sns.violinplot(x = 'Source', y = 'Number of PTMs', hue = 'UniProt Isoform Type', palette = ['gray', palette[1]], data = self.isoform_ptms, ax = ax)
        ax.legend(bbox_to_anchor = (1.1, 1.25), fontsize = 9)
        ax.set_xlabel('')

        if panel_save_dir is not None:
            plt.savefig(panel_save_dir + 'SFig2B.svg', bbox_inches ='tight', dpi = 300)

    def PanelC(self, ax = None, panel_save_dir = None):
        #pre mapping
        ptm_info = self.data.mapper.ptm_info.copy()
        pre_canonical = ptm_info[ptm_info['Isoform Type'] == 'Canonical'].shape[0]
        pre_alternative = ptm_info[ptm_info['Isoform Type'] == 'Alternative'].shape[0]

        #post mapping
        identified = self.data.mapper.isoform_ptms[self.data.mapper.isoform_ptms["Mapping Result"] == 'Success']
        post_canonical = identified[identified['Isoform Type'] == 'Canonical'].shape[0]
        post_alternative = identified[identified['Isoform Type'] == 'Alternative']
        post_alt_ensembl = post_alternative[post_alternative['Isoform ID'].str.contains('ENS')]
        post_alt_uniprot = post_alternative[~post_alternative['Isoform ID'].str.contains('ENS')]

        canonical = [post_canonical, pre_canonical]
        alt_uniprot = [post_alt_uniprot.shape[0], pre_alternative]
        alt_ensembl = [post_alt_ensembl.shape[0],0]

        #plot stacked bar plot
        fig, ax = plt.subplots(figsize = (2.5,2))
        palette = sns.color_palette('colorblind')
        ax.barh(['Post-Mapping', 'Pre-Mapping'], canonical, color = 'gray', label = 'Canonical')
        ax.barh(['Post-Mapping', 'Pre-Mapping'], alt_uniprot, left = canonical, color = palette[1], label = 'Alternative (UniProtKB)')
        ax.barh(['Post-Mapping', 'Pre-Mapping'], alt_ensembl, left = np.array(canonical) + np.array(alt_uniprot), color = 'bisque', label = 'Alternative (Ensembl Only)')
        ax.legend(bbox_to_anchor = (0, 1.5), loc = 'upper left', fontsize = 9)
        ax.set_xlabel('Number of PTMs across all Isoforms', labelpad = 14)

        #add line indicating fraction of PTMs that were newly identified
        frac_new = 1- pre_alternative/(post_alt_ensembl.shape[0] + post_alt_uniprot.shape[0])
        ax.plot([pre_canonical + pre_alternative, post_canonical + post_alt_uniprot.shape[0] + post_alt_ensembl.shape[0]], [0.5, 0.5], color = 'black', linewidth = 1)
        ax.text(pre_canonical + (post_alt_uniprot.shape[0]+post_alt_ensembl.shape[0])/2, 0.5, f"{round(frac_new*100,1)}%\nnot annotated", va = 'bottom', ha = 'center', fontsize = 9)


        if panel_save_dir is not None:
            plt.savefig(panel_save_dir + 'SFig2C.svg', bbox_inches ='tight', dpi = 300)




    def generate_figure(self, figsize = (6.5, 3.5), fig_save_dir = None):
        """
        generate supplementary figure 2
        """
        fig, axes = plt.subplots(ncols = 2, figsize = figsize, gridspec_kw = {'width_ratios':[1,1.5], 'wspace':0.5})

        #Panel A
        self.PanelA(axes[0])
        fig.text(0.01, 0.85, 'A', fontsize = 24, fontweight = 'bold')
        #Panel B
        self.PanelB(axes[1])
        fig.text(0.4, 0.85, 'B', fontsize = 24, fontweight = 'bold')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'SupplementaryFigure2.pdf', bbox_inches = 'tight')




class SupplementaryFigure4:
    """
    PTM density in exons
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data

        ###### extra exon specific information ########
        #grab canonical transcript information
        canonical_transcripts = config.translator.loc[config.translator['UniProt Isoform Type'] == 'Canonical', 'Transcript stable ID'].unique()

        #grab exons associated with transcripts
        trim_exons = self.data.mapper.exons[self.data.mapper.exons['Transcript stable ID'].isin(canonical_transcripts)]
        #restrict to exons with coding information, remove rows with missing info
        trim_exons = trim_exons.loc[trim_exons['Warnings (AA Seq)'].isna(), ['Exon stable ID', 'Exon AA Seq (Full Codon)', 'Constitutive exon']]
        trim_exons = trim_exons.dropna()
        #calculate the exon length in amino acids
        trim_exons['Number of AA'] = trim_exons['Exon AA Seq (Full Codon)'].apply(len)
        trim_exons = trim_exons.drop_duplicates()
        #remove exons with conflicting info
        trim_exons = trim_exons[~trim_exons.duplicated(subset = 'Exon stable ID', keep = False)]

        # calculate the number of ptms in each exon from exploded_ptm data
        tmp_ptms = self.data.exploded_ptms.drop_duplicates(['PTM', 'Exon stable ID']).copy()
        num_ptms = tmp_ptms.groupby(['Protein','Exon stable ID']).size().reset_index()
        num_ptms = num_ptms[~num_ptms.duplicated(subset = 'Exon stable ID', keep = False)]
        trim_exons = trim_exons.merge(num_ptms, on = 'Exon stable ID', how = 'left')
        trim_exons[0] = trim_exons[0].replace(np.nan, 0)
        #calculate density of PTMs (Number of PTMs/number of residues)
        trim_exons['Density'] = (trim_exons[0]/trim_exons['Number of AA'])

        self.trim_exons = trim_exons

    def PanelA(self, trim_exons, ax = None):
        """
        Given trim exon storing PTM density in 'Density' column, plot violin plot comparing constitutive vs. non-constitutive exons

        Parameters
        ----------
        trim_exons : pandas dataframe
            Dataframe storing PTM density in 'Density' column
        ax : matplotlib axis object
            Axis to plot on
        fig_save_dir : str, optional
            Directory to save figure to. The default is None and will not save figure.
        data_save_dir : str, optional
            Directory to save data to. The default is None and will not save source data.
        """
        if ax is None:
            fig, ax = plt.subplots(figsize = (3,3))
        
        sns.violinplot(y = 'Density', x = 'Constitutive exon', data = trim_exons, ax = ax)
        ax.set_xticklabels(['Non-constitutive', 'Constitutive'], fontsize = 9)
        ax.set_ylabel('PTM Density')
        ax.set_xlabel('')
        ax.set_ylim([0,1.2])
        ax.set_yticks([0,0.2,0.4,0.6,0.8,1])

        #add stat annotation
        f, p = stats.mannwhitneyu(trim_exons.loc[trim_exons['Constitutive exon'] == 0, 'Density'], trim_exons.loc[trim_exons['Constitutive exon'] == 1, 'Density'])
        plot_types.addStatAnnot(p, 0, 1, ax, h = 0.02, col = 'k', equal_var = False, start_height =1.1)


    def PanelB(self, trim_exons, gs = None, fig = None):
        """
        Make SFigure 4B, which shows a 2D histogram of PTM density vs. exon length for constitutive and non-constitutive exons
        """
        const_exons = trim_exons.loc[trim_exons['Constitutive exon'] == 1]
        nonconst_exons = trim_exons.loc[trim_exons['Constitutive exon'] == 0]

        if gs is None:
            fig, ax = plt.subplots(ncols = 2, figsize = (5,3), sharey = True)
            fig.subplots_adjust(wspace = 0.6)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs, wspace = 0.4)
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1])]
        #plot 2D histograms
        hist1 = ax[0].hist2d(np.log2(nonconst_exons['Number of AA']), nonconst_exons['Density'], bins = 20, norm = 'symlog', cmap = 'Blues', vmax = 1000)
        ax[0].set_xlabel('Log2(Exon Length)')
        ax[0].set_ylabel('PTM Density')
        ax[0].set_ylim([0,1])
        hist2 = ax[1].hist2d(np.log2(const_exons['Number of AA']), const_exons['Density'], bins = 20, norm = 'log', cmap = 'Oranges', vmax = 1000)
        ax[1].set_xlabel('Log2(Exon Length)')
        ax[1].set_ylim([0,1])
        #ax[1].set_yticklabels([])

        #plot colorbar for non-constitutive plot (first histogram)
        cbar_ax = fig.add_axes([0.49, 1.1, 0.18, 0.02])
        cbar = fig.colorbar(hist1[3], cax=cbar_ax, orientation = 'horizontal')
        cbar.set_label('Number of\nNon-Constitutive Exons', fontsize = 9)


        #plot colorbar for constitutive plot
        cbar_ax = fig.add_axes([0.73, 1.1, 0.18, 0.02])
        cbar = fig.colorbar(hist2[3], cax=cbar_ax, orientation = 'horizontal')
        cbar.set_label('Number of\nConstitutive Exons', fontsize = 9)


        return ax
    
    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        #set up figure
        fig = plt.figure(figsize = (6.5, 3))
        gs = gridspec.GridSpec(1, 2, width_ratios=[0.7, 1], wspace = 0.3)

        #make figure4A
        ax = fig.add_subplot(gs[0])
        self.PanelA(self.trim_exons, ax = ax)
        ax.text(-1, 1.3, 'A', fontsize = 24, weight = 'bold')

        #make figure4B
        ax = self.PanelB(self.trim_exons, gs = gs[1], fig = fig)
        ax[0].text(-7, 1.1, 'B', fontsize = 24, weight = 'bold')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure4.{fig_type}', format = fig_type, bbox_inches = 'tight', dpi = 300)



    

class SupplementaryFigure5:
    """
    Supplementary Figure 5, which displays the constitutive rates for all classes of modifications. Also displays the number of instances of each class, and sorts by this number
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data
        #get constitutive rates for broad classes
        self.constitutive_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)


    def generate_figure(self, ax = None, fig_save_dir = None, fig_type = 'svg'):
        if ax is None:
            fig, ax = plt.subplots(figsize = (6,8), ncols = 2, sharey = True, width_ratios = [1, 0.3])
            fig.subplots_adjust(wspace = 0)




        plt_data = self.constitutive_rates.sort_values(by = 'Number of Instances in Proteome', ascending = True).copy()
        #convert modification names to recognizable names
        plt_data = plt_data.rename(self.data.class_name_conversion)

        #pick a colorpalette, create bar plot
        colors = sns.color_palette('terrain', plt_data.shape[0])[::-1] 
        ax[0].barh(plt_data.index, plt_data['Rate'].values, color =colors)

        #adjust plot parameters
        ax[0].set_xticks([0, 0.2, 0.4, 0.6, 0.8])
        ax[0].set_xlim([0,1])
        ax[0].set_xlabel('Constitutive PTM Rate')
        xticks = ax[0].get_yticks()
        ax[1].set_xticks([])
        ax[1].set_xlabel('Number\nof Instances', rotation = 0, ha = 'center')
        ax[1].set_ylim([-0.5, plt_data.shape[0]-0.5])

        #add number of modification types being profiled
        for tick in xticks:
            ax[1].annotate(plt_data.iloc[int(tick),0],(0.5,tick),ha = 'center', va = 'center', fontsize = 9)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure5.{fig_type}', bbox_inches = 'tight')



class SupplementaryFigure6:
    """
    Constitiutive rates for all modification classes
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data
        self.constitutive_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)

    def PanelA(self, ax = None, ylim = [0.5,0.8]):
        """
        Make Supplementary Figure 6A, which shows the constitutive rate of PTMs within and outside of domains

        Parameters
        ----------
        ax : matplotlib axis object
            Axis to plot on
        data_save_dir : str
            Directory to save source data, if None, source data will not be saved
        fig_save_dir : str
            Directory to save figure, if None, figure will not be saved

        Returns
        -------
        None
        """
        if ax is None:
            fig, ax = plt.subplots(figsize = (2,4))

        #grab boolean that indicates if ptm is constitutive
        isConstitutive = self.data.mapper.ptm_info['PTM Conservation Score'] == 1
        #grab boolean that indicates if ptm is within a domain
        inDomain = self.data.mapper.ptm_info['inDomain']

        #calculate constitutive rates for ptms within and outside of domains
        inDomain_rate= self.data.mapper.ptm_info[isConstitutive & inDomain].shape[0]/self.data.mapper.ptm_info[inDomain].shape[0]
        outDomain_rate = self.data.mapper.ptm_info[isConstitutive & ~inDomain].shape[0]/self.data.mapper.ptm_info[~inDomain].shape[0]

        #plot barplot
        ax.bar(['Yes', 'No'], [inDomain_rate, outDomain_rate])
        ax.set_xlabel('Within Domain?')
        ax.set_ylabel('Constitutive Rate')
        ax.set_ylim(ylim)

    def PanelB(self, ax = None):
        """
        Scatterplot showing relationship between constitutive rate and fraction of ptms in domains
        """
        #get indomain rates for each modification class
        tmp = self.data.exploded_mods.groupby('Modification Class')['inDomain'].sum()/self.data.exploded_mods.groupby('Modification Class').size()

        #combine constitutive rates with indomain rates
        plt_data = pd.concat([self.constitutive_rates,tmp], axis = 1)
        plt_data = plt_data.rename(columns = {0:'Within Domain Rate'})

        #restrict to mods with at least 150 sites
        plt_data = plt_data[plt_data['Number of Instances in Proteome'] >= self.data.min_mods]

        #plot scatterplot
        plt.scatter(plt_data['Within Domain Rate'], plt_data['Rate'])
        plt.xlabel('Within Domain Rate')
        plt.ylabel('Constitutive Rate')

        label_loc = {'Phosphorylation': (-0.002,0.002,'right'),
                    'Ubiquitination':(0,0,'left'),
                    'Acetylation':(0,0.005,'left'),
                    'Methylation':(0,0,'right'),
                    'Hydroxylation':(0,0,'right'),
                    'Sumoylation':(0,-0.01,'left'),
                    'Dimethylation':(-0.02,-0.01, 'center'),
                    'Trimethylation':(0, -0.005, 'left'),
                    'Succinylation':(0,0.005,'right')}
        #annotate the plot with the modification class
        for i, txt in enumerate(plt_data.index):
            if txt in label_loc:
                loc = (plt_data['Within Domain Rate'].iloc[i]+label_loc[txt][0], plt_data['Rate'].iloc[i]+label_loc[txt][1])
                orientation = label_loc[txt][2]
            else:
                loc = (plt_data['Within Domain Rate'].iloc[i], plt_data['Rate'].iloc[i])
                orientation = 'left'

            plt.annotate(txt, loc, ha = orientation, fontsize = 8)

        #remove top and right spines
        plt.gca().spines['top'].set_visible(False)
        plt.gca().spines['right'].set_visible(False)


        #perform correlation analysis
        r,p = stats.pearsonr(plt_data['Within Domain Rate'], plt_data['Rate'])
        #annotate upper right hand portion of plot with correlation coefficient and significance
        plt.text(0.8, 0.95, 'r = {}, p = {}'.format(round(r,2), round(p,2)), horizontalalignment='center', verticalalignment='center', transform = plt.gca().transAxes)



    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        """
        Generate supplementary figure 6
        """


        fig, axes = plt.subplots(ncols = 2, figsize = (8.5, 4.5), gridspec_kw={'width_ratios': [1, 4], 'wspace': 0.3})
        self.PanelA(ax = axes[0])
        fig.text(0.03, 0.85, 'A', fontsize = 24, fontweight = 'bold')
        self.PanelB(ax = axes[1])
        fig.text(0.28, 0.85, 'B', fontsize = 24, fontweight = 'bold')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure6.{fig_type}', format = fig_type, bbox_inches = 'tight', dpi = 300)






class SupplementaryFigure7:
    """
    Constitutive rates for modification subtypes
    """
    def __init__(self, data):
        """
        Subtypes
        """
        #point to figure data class
        self.data = data

        #get subtype rates for individual modification types
        self.class_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)
        self.subtype_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationSubtype.csv', index_col = 0)

        #get classes to plot
        self.class_to_analyze, self.trim_mod_groups = self.identifyClassesToAnalyze()
        self.num_subtypes = self.trim_mod_groups.groupby('Modification Class')['Modification'].nunique() 
        self.num_classes = self.trim_mod_groups['Modification Class'].nunique()

        #establish color palette
        self.colors = sns.color_palette('terrain', self.num_classes)

    def identifyClassesToAnalyze(self, min_subtypes = 1):
        """
        Get modification classes with at least <min_subtypes>, and sort by number of modifications
        """
        trim_mod_groups = self.data.mod_groups[self.data.mod_groups['Modification'].isin(self.subtype_rates.index)]
        num_subtypes = trim_mod_groups.groupby('Modification Class')['Modification'].nunique() 
        #get classes with subtypes
        class_to_analyze = [sub for sub in num_subtypes[num_subtypes > min_subtypes].index if sub in trim_mod_groups['Modification Class'].values]
        #sort classes by number of modifications
        class_to_analyze = self.class_rates.loc[class_to_analyze].sort_values(by = 'Number of Instances in Proteome', ascending = False).index
        return class_to_analyze, trim_mod_groups

    def panels(self, class_to_analyze, color, gs = None, fig = None, fontsize = 9):
        """
        Plot a plot with the constitutive rate for each modification subtype
        """
        modifications = self.trim_mod_groups.loc[self.trim_mod_groups['Modification Class'] == class_to_analyze, 'Modification']
        trim_rates = self.subtype_rates.loc[modifications]
        trim_rates = trim_rates.sort_values(by = 'Number of Instances in Proteome', ascending = False)
        #rate plot
        if gs is None:
            fig, ax = plt.subplots(figsize = (trim_rates.shape[0]/3+1.1**trim_rates.shape[0], 2.5), nrows = 2, height_ratios = [1, 0.15])
            fig.subplots_adjust(hspace = 0)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(2, 1, subplot_spec=gs, height_ratios = [1, 0.15], hspace = 0)
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1])]
        ax[0].bar(trim_rates.index, trim_rates['Rate'], color = color)
        ax[0].set_ylim([0,1])
        ax[0].set_ylabel('Constitutive\nRate', fontsize = fontsize)
        
        xticks = ax[0].get_xticks()
        xtick_labels = ax[0].get_xticklabels()
        ax[0].set_xticks([])
        ax[1].set_yticks([])
        ax[1].set_ylabel('Number of PTM Type', rotation = 0, ha = 'right', va = 'top', fontsize = fontsize)
        ax[1].set_xlim([-0.5, trim_rates.shape[0]-0.5])
        ax[0].set_xlim([-0.5, trim_rates.shape[0]-0.5])
        ax[0].set_title(class_to_analyze)
        for tick in xticks:
            ax[1].annotate(trim_rates.iloc[tick,0],(tick, 0.8),ha = 'center', va = 'top', fontsize = 9)
        ax[1].set_xticks(xticks)
        ax[1].set_xticklabels(xtick_labels, rotation = 35, ha = 'right', fontsize = fontsize)

    def generate_page1(self, fig_save_dir = None, fig_type = 'svg'):
        """
        Plot subtypes for phosphorylation, glycosylation, acetylation, and methylation
        """
        fig = plt.figure(figsize = (6.5, 7.5))
        gs = gridspec.GridSpec(3, 1,hspace = 1)
        gs_row1 = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[0], wspace = 0.8, width_ratios=[1,0.5])

        #PHOS
        self.panels('Phosphorylation', gs = gs_row1[0], fig = fig, color = self.colors[0])
        fig.text(0.02, 0.9, 'A', weight = 'bold', fontsize = 24)


        #GLCN
        self.panels('Glycosylation', gs = gs_row1[1], fig = fig, color = self.colors[1])
        fig.text(0.5, 0.9, 'B', weight = 'bold', fontsize = 24)

        #ACET
        self.panels('Acetylation', gs = gs[1], fig = fig, color = self.colors[2])
        fig.text(0.02, 0.58, 'C', weight = 'bold', fontsize = 24)

        #METH
        self.panels('Methylation', gs = gs[2], fig = fig, color = self.colors[3])
        fig.text(0.02, 0.28, 'D', weight = 'bold', fontsize = 24)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure7_Page1.{fig_type}', bbox_inches = 'tight')

    def generate_page2(self, fig_save_dir = None, fig_type = 'svg'):
        """
        Plot subtypes for dimethylation, trimethylation, hydroxylation, sulfation, and myristolyation
        """
        fig = plt.figure(figsize = (6.5, 7.5))
        gs = gridspec.GridSpec(3, 1,hspace = 1.5)
        gs_row1 = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[0], wspace = 0.8, width_ratios=[1,0.5])
        gs_row3 = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[2], wspace = 2, width_ratios=[1,1])

        #DIMETH
        self.panels('Dimethylation', gs = gs_row1[0], fig = fig, color = self.colors[4])
        fig.text(0.02, 0.9, 'E', weight = 'bold', fontsize = 24)


        #TRIMETH
        self.panels('Trimethylation', gs = gs_row1[1], fig = fig, color = self.colors[5])
        fig.text(0.5, 0.9, 'F', weight = 'bold', fontsize = 24)

        #HYRD
        self.panels('Hydroxylation', gs = gs[1], fig = fig, color = self.colors[6])
        fig.text(0.02, 0.58, 'G', weight = 'bold', fontsize = 24)

        #SULF
        self.panels('Sulfation', gs = gs_row3[0], fig = fig, color = self.colors[7])
        fig.text(0.02, 0.28, 'H', weight = 'bold', fontsize = 24)

        #MYRI
        self.panels('Myristoylation', gs = gs_row3[1], fig = fig, color = self.colors[8])
        fig.text(0.62, 0.28, 'I', weight = 'bold', fontsize = 24)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure7_Page2.{fig_type}', bbox_inches = 'tight')

    def generate_page3(self, fig_save_dir = None, fig_type = 'svg'):
        """
        Plot subtypes for amidation, ADP-ribosylation, and deamidation
        """
        fig = plt.figure(figsize = (6.5, 5))
        gs = gridspec.GridSpec(2, 1,hspace = 1.5)
        gs_row2 = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[1], wspace = 1, width_ratios=[1,0.5])

        #AMID
        self.panels('Amidation', gs = gs[0], fig = fig, color = self.colors[9])
        fig.text(0.02, 0.9, 'J', weight = 'bold', fontsize = 24)


        #ADP
        self.panels('ADP Ribosylation', gs = gs_row2[0], fig = fig, color = self.colors[10])
        fig.text(0.02, 0.35, 'K', weight = 'bold', fontsize = 24)

        #DEAM
        self.panels('Deamidation', gs = gs_row2[1], fig = fig, color = self.colors[11])
        fig.text(0.6, 0.35, 'L', weight = 'bold', fontsize = 24)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure7_Page3.{fig_type}', bbox_inches = 'tight')

    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        self.generate_page1(fig_save_dir=fig_save_dir, fig_type = fig_type)
        self.generate_page2(fig_save_dir=fig_save_dir, fig_type = fig_type)
        self.generate_page3(fig_save_dir=fig_save_dir,  fig_type = fig_type)






class SupplementaryFigure8:
    """
    Constitituve rates using different functional filters
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data
        #load rates
        self.constitutive_rates = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/ByModificationClass.csv', index_col = 0)

        #load data
        self.database_filter = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Filtered/ByDatabase.csv', index_col = 0)
        self.tsl_filter = pd.read_csv(self.data.analysis_dir+'/Constitutive_Rates/Filtered/ByTranscriptSupportLevel.csv', index_col = 0)
        self.trifid_filter = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Filtered/ByTRIFID.csv', index_col = 0)

        #load mapper object with additional constitutive rate scores
        self.data.mapper.ptm_info = pd.read_csv(self.data.ptm_data_dir + '/processed_data_dir/ptm_info.csv', index_col = 0)

    def PanelA(self, gs = None, fig = None):
        """
        Filtering transcripts based on whether they appear in a specific database
        """
        if gs is None:
            fig, ax = plt.subplots(figsize = (2.5,4), nrows = 2, sharex = True) 
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(2, 1, subplot_spec=gs)
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1])]

        #plot barplots
        ax[0].bar(self.database_filter.index, self.database_filter['Constitutive Rate'], color = 'gray')
        ticks = plt.xticks(rotation = 90)
        ax[0].set_ylabel('Constitutive\nRate', fontsize = self.data.axes_label_size)
        ax[0].set_yticks([0,0.2,0.4,0.6,0.8,1.0])
        if gs is not None:
            ax[0].set_xticklabels([])

        ax[1].bar(self.database_filter.index, np.array(list(self.database_filter['Number of Isoforms'].values))/self.data.mapper.ptm_info['Protein'].nunique(), color = 'gray')
        ax[1].set_ylabel('# of Isoforms\nPer Protein', fontsize = self.data.axes_label_size)
        ax[1].set_ylim([0, 2.5])

    def PanelB(self, gs = None, fig = None):
        """
        Plot the constitutive rate for each transcript support level
        """
        if gs is None:
            fig, ax = plt.subplots(figsize = (2.5,4), nrows = 2, sharex = True)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(2, 1, subplot_spec=gs)
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1])]

        ax[0].bar(self.tsl_filter.index, self.tsl_filter['Constitutive Rate'], color = 'gray')
        ax[0].set_ylabel('Constitutive\nRate', fontsize = self.data.axes_label_size)
        ax[0].set_yticks([0,0.2,0.4,0.6,0.8,1.0])
        if gs is not None:
            ax[0].set_xticklabels([])
        ax[1].bar(self.tsl_filter.index, np.array(list(self.tsl_filter['Number of Isoforms'].values))/self.data.mapper.ptm_info['Protein'].nunique(), color = 'gray')
        ax[1].set_ylabel('# of Isoforms\nPer Protein', fontsize = self.data.axes_label_size)
        ax[1].set_ylim([0, 2.5])
        ax[1].set_xlabel('Transcript support Level', fontsize = self.data.axes_label_size)
        ticks = ax[1].set_xticklabels(['Any', '$\geq\!5$', '$\geq\!4$', '$\geq\!3$', '$\geq\!2$', '1'], fontsize = self.data.axes_label_size - 1)


    def PanelC(self, ax = None):
        """
        Constitutive rates for different TRIFID thresholds
        """
        if ax is None:
            fig, ax = plt.subplots(figsize = (2.5,4))

        palette = sns.color_palette('colorblind')
        ax_color = palette[0]
        ax.plot(self.trifid_filter.index, self.trifid_filter['Constitutive Rate'].values, color = ax_color, marker = 'o', ms = 3)
        ax.set_ylabel('Constitutive Rate', color = ax_color, fontsize = self.data.axes_label_size-1)
        ax.tick_params(axis='y', labelcolor = ax_color)
        ax.set_ylim([0,1])
        ax2 = ax.twinx()
        ax2_color = palette[1]
        ax2.plot(self.trifid_filter.index, np.array(list(self.trifid_filter['Number of Isoforms'].values))/self.data.mapper.ptm_info['Protein'].nunique(), color = ax2_color, marker = 'o', ms = 3)
        ax2.set_ylabel('# of Isoforms Per Protein', rotation = 270, va = 'bottom', c = ax2_color, fontsize = self.data.axes_label_size)
        ax2.tick_params(axis='y', labelcolor = ax2_color)
        ax2.set_ylim([0,2])
        ax.set_xlabel('TRIFID Threshold Value')
        ax.set_xlim([-0.01,1])


    def PanelD(self, gs = None, fig = None, min_mods = 150):
        """
        Plot how constitutive rate for specific modification classes are when filtered by databases
        """
        #construct figure
        if gs is None:
            fig, ax = plt.subplots(figsize = (8,4), nrows = 3, sharex = True)
            fig.subplots_adjust(hspace = 0.5)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(3, 1, subplot_spec=gs, hspace = 0.5)
            ax = [fig.add_subplot(gs2[0]), fig.add_subplot(gs2[1]), fig.add_subplot(gs2[2])]

        
        rates = self.constitutive_rates.sort_values(by = 'Number of Instances in Proteome', ascending = False).copy()
        mods_to_keep = rates[rates['Number of Instances in Proteome'] >= min_mods]
        mods_to_keep = mods_to_keep.sort_values(by = 'Rate').index.values
        #labels = []
        #for mod in mods_to_keep:
        #    labels.append(self.data.class_name_conversion[mod])
            
        ax[0].bar(mods_to_keep, rates.loc[mods_to_keep, 'Rate'], color = 'gray')
        ax[0].set_ylim([0.4,0.9])
        ax[0].set_yticks([0.4, 0.5, 0.6, 0.7, 0.8,0.9])
        if gs is not None:
            ax[0].set_xticklabels([])
        crate = self.data.mapper.ptm_info[self.data.mapper.ptm_info['PTM Conservation Score'] == 1].shape[0]/self.data.mapper.ptm_info.shape[0]
        ax[0].axhline(crate, linestyle = 'dashed', c = 'black', alpha = 0.5)
        ax[0].set_title('Ensembl', fontsize = self.data.axes_label_size)

        ## UniProt (all)
        fully_conserved = self.data.mapper.ptm_info[self.data.mapper.ptm_info['UniProt Isoform Score'] == 1].copy()
        fully_conserved["Modification Class"] = fully_conserved['Modification Class'].apply(lambda x: x.split(';'))
        fully_conserved = fully_conserved.explode('Modification Class').reset_index()
        fully_conserved = fully_conserved.rename({'index':'PTM'}, axis = 1)
        #fully_conserved = fully_conserved.merge(self.data.mod_groups[['Mod Name', 'Mod Class']], left_on = 'Modification', right_on = 'Mod Name')
        #fully_conserved = fully_conserved.drop_duplicates()
        #get constitutive ptms, then add in any mod types that don't have any constitutive ptms
        grouped_conserved = fully_conserved.groupby('Modification Class').size()
        for mod in mods_to_keep:
            if mod not in grouped_conserved.index.values:
                grouped_conserved[mod] = 0

        plt_data_uniprot = grouped_conserved[mods_to_keep]/self.data.exploded_mods.groupby('Modification Class').size()[mods_to_keep]
        #pick a colorpalette, create bar plot
        plt_data_uniprot = plt_data_uniprot[mods_to_keep]
        ax[2].bar(mods_to_keep, plt_data_uniprot.values, color ='gray')

        #get and plot overall constitutive rate
        crate = self.data.mapper.ptm_info[self.data.mapper.ptm_info['UniProt Isoform Score'] == 1].shape[0]/self.data.mapper.ptm_info.shape[0]
        ax[2].axhline(crate, alpha = 0.5, c = 'black', linestyle = 'dashed')
        #adjust ticks
        ax[2].set_yticks([0.7, 0.8, 0.9, 1])
        ax[2].set_ylim([0.7,1])
        ax[2].set_title('UniProtKB', fontsize = self.data.axes_label_size)

        ## APPRIS
        fully_conserved = self.data.mapper.ptm_info[self.data.mapper.ptm_info['APPRIS Score'] == 1].copy()
        fully_conserved["Modification Class"] = fully_conserved['Modification Class'].apply(lambda x: x.split(';'))
        fully_conserved = fully_conserved.explode('Modification Class').reset_index()
        fully_conserved = fully_conserved.rename({'index':'PTM'}, axis = 1)
        #fully_conserved = fully_conserved.merge(self.data.mod_groups[['Mod Name', 'Mod Class']], left_on = 'Modification', right_on = 'Mod Name')
        #fully_conserved = fully_conserved.drop_duplicates()
        #get constitutive ptms, then add in any mod types that don't have any constitutive ptms
        grouped_conserved = fully_conserved.groupby('Modification Class').size()

        for mod in mods_to_keep:
            if mod not in grouped_conserved.index.values:
                grouped_conserved[mod] = 0

        plt_data_appris = grouped_conserved[mods_to_keep]/self.data.exploded_mods.groupby('Modification Class').size()[mods_to_keep]
        #pick a colorpalette, create bar plot
        plt_data_appris = plt_data_appris[mods_to_keep]
        ax[1].bar(mods_to_keep, plt_data_appris.values, color ='gray')

        #get and plot overall constitutive rate
        crate = self.data.mapper.ptm_info[self.data.mapper.ptm_info['APPRIS Score'] == 1].shape[0]/self.data.mapper.ptm_info.shape[0]
        ax[1].axhline(crate, alpha = 0.5, c = 'black', linestyle = 'dashed')
        #adjust ticks
        ticks = plt.xticks(rotation = 35, ha = 'right', fontsize = self.data.axes_label_size-1)

        #ax[2].set_yticks([0, 0.25, 0.5, 0.75, 1])
        ax[1].set_ylim([0.9,1])
        ax[1].set_title('APPRIS', fontsize = self.data.axes_label_size)
        ax[1].set_ylabel('Constitutive Rate', fontsize = self.data.axes_label_size)
        if gs is not None:
            ax[1].set_xticklabels([])



    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1], hspace = 0.5)
        gs_row1 = gridspec.GridSpecFromSubplotSpec(1, 3, wspace = 0.6,width_ratios = [1,1,1], subplot_spec=gs[0])
        

        fig = plt.figure(figsize = (6.5, 7))
        #Figure S9A
        self.PanelA(gs = gs_row1[0], fig = fig)
        fig.text(0.01, 0.95, 'A', fontsize=24, fontweight='bold', va='top', ha='left')

        #Figure S9B
        self.PanelB(gs = gs_row1[1], fig = fig)
        fig.text(0.3, 0.95, 'B', fontsize=24, fontweight='bold', va='top', ha='left')

        #Figure S9C
        ax = fig.add_subplot(gs_row1[2])
        self.PanelC(ax = ax)
        fig.text(0.6, 0.95, 'C', fontsize=24, fontweight='bold', va='top', ha='left')

        #Figure S9D
        self.PanelD(gs = gs[1], fig = fig)
        fig.text(0.01, 0.45, 'D', fontsize=24, fontweight='bold', va='top', ha='left')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure8.{fig_type}', bbox_inches = 'tight', dpi = 300)




class SupplementaryFigure9:
    """
    Functional enrichment of constitutive and non-constitutive ptms
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data

        #load enrichment data
        self.enrichment_results = {'Function':pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Function_Enrichment.csv'), 'Process':pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Process_Enrichment.csv'), 'Kinase':pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Kinase_Enrichment.csv')}


    def plot_log_odds(self, ax = None, annotation_column = 'Function'):
        if ax is None:
            fig, ax = plt.subplots(figsize = (3,4))

        plt_data = self.enrichment_results[annotation_column].copy()
        plt_data = plt_data[plt_data['BH adjusted p'] <= 0.05]
        plt_data = plt_data.sort_values(by = 'Log Odds', ascending = True)

        #replace inf with max odds ratio
        plt_data['Log Odds'] = plt_data['Log Odds'].replace([np.inf, -np.inf], np.nan)
        plt_data['Log Odds'] = plt_data['Log Odds'].fillna(plt_data['Log Odds'].max())

        #color bar between non-constitutive and constitutive
        colors = []
        for i, row in plt_data.iterrows():
            if row['Log Odds'] >= 0:
                colors.append('deepskyblue')
            else:
                colors.append('gold')

        #plot horizontal barplot with log p values
        ax.barh(plt_data[annotation_column], plt_data['Log Odds'], color = colors, edgecolor = 'black')
        ax.set_xticks([-2,0,2])
        ax.set_xlabel('Log2(Odds Ratio)', fontsize = self.data.axes_label_size-1)
        ax.axvline(0, color = 'black', linestyle = '-')
        ax.set_ylim(-0.5, len(plt_data[annotation_column])-0.5)
        ax.tick_params(labelsize = self.data.axes_label_size-1)

        #turn off left, right, and to spines
        ax.spines['left'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)

    def PanelA(self, ax = None):
        self.plot_log_odds(ax = ax, annotation_column = 'Function')

    def PanelB(self, ax = None):
        self.plot_log_odds(ax = ax, annotation_column='Process')
    
    def PanelC(self, ax = None):
        self.plot_log_odds(ax = ax, annotation_column='Kinase')


    def legend(self, ax = None):
        if ax is None:
            fig, ax = plt.subplots(figsize = (1,0.3))

        ax.axis('off')
        legend_handles = [Patch(facecolor = 'gold', label = 'Non-constitutive'), Patch(facecolor = 'deepskyblue', label = 'Constitutive')]
        ax.legend(handles = legend_handles, ncols = 2, title = 'Enriched in:', loc = 'center', fontsize = self.data.axes_label_size)   

    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        gs = gridspec.GridSpec(2, 1, height_ratios=[0.95, 0.05], hspace = 0.4)

        fig = plt.figure(figsize = (6.5, 5))
        #plot legend
        ax0 = fig.add_subplot(gs[1])
        self.legend(ax = ax0)

        gs_figs = gridspec.GridSpecFromSubplotSpec(1, 3, width_ratios=[1, 1,1.2], wspace = 1.5, subplot_spec=gs[0])
        #Panel A
        ax10 = fig.add_subplot(gs_figs[0])
        self.PanelA(ax = ax10)
        fig.text(-0.1, 0.93, 'A', fontsize=24, fontweight='bold', va='top')

        #Panel B
        ax11 = fig.add_subplot(gs_figs[1])
        self.PanelB(ax = ax11)
        fig.text(0.28, 0.93, 'B', fontsize=24, fontweight='bold', va='top')

        #Panel C
        ax12 = fig.add_subplot(gs_figs[2])
        self.PanelC(ax = ax12)
        fig.text(0.6, 0.93, 'C', fontsize=24, fontweight='bold', va='top')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure9.{fig_type}', bbox_inches = 'tight', dpi = 300)




class SupplementaryFigure10:
    """
    Mutually exclusive exon event information
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data

        mxe_ptms = self.data.mapper.alternative_ptms[self.data.mapper.alternative_ptms['Event Type'] == 'Mutually Exclusive'].copy()
        exploded_mxes = mxe_ptms.copy()
        exploded_mxes['Modification'] = exploded_mxes["Modification"].apply(lambda x: x.split(';'))
        exploded_mxes = exploded_mxes.explode('Modification').drop_duplicates(subset = ['Source of PTM', 'Exon ID (Canonical)', 'Exon ID (Alternative)'])
        exploded_mxes['Match'] = exploded_mxes['Mapping Result'] == 'Success'
        self.mxe_ptms = mxe_ptms
        self.exploded_mxes = exploded_mxes

    def PanelB(self, ax = None, panel_save_dir = None):
        """
        Breakdown of PTM types involved in MXE events
        """
        mod_sizes = self.exploded_mxes.groupby('Modification').size().sort_values(ascending = False)
        num_matched = self.exploded_mxes.groupby('Modification')['Match'].sum()
        fractions = num_matched[mod_sizes.index]/mod_sizes
        if ax is None:
            fig, ax = plt.subplots(figsize=(3, 3))

        ax.bar(fractions.index, fractions.values, color = 'grey')

        ax.set_xticklabels(ax.get_xticklabels(), rotation = 35, ha = 'right', fontsize = self.data.axes_label_size-1)
        ax.set_ylabel('Fraction of MXEs\nwith Conserved PTM', fontsize = self.data.axes_label_size)
        ticks = ax.get_xticks()
        for i, size in enumerate(mod_sizes):
            if fractions.iloc[i] == 1:
                ax.annotate(size, (ticks[i], 0.9), ha = 'center', fontsize = self.data.axes_label_size)
            else:
                ax.annotate(size, (ticks[i], fractions.iloc[i]+0.01), ha = 'center', fontsize = self.data.axes_label_size)
        ax.set_ylim([0,1])
        ax.set_xlim([-0.5, 12.5])
        ax.set_yticklabels(ax.get_yticklabels(), fontsize = self.data.axes_label_size-1)

        if panel_save_dir is not None:
            plt.savefig(panel_save_dir + 'SFig10B_MXE_Mods.svg', bbox_inches = 'tight')

    def PanelC(self, min_instances = 5, gs = None, fig = None, panel_save_dir = None):
        """
        Indicate which positions are altered in cases where MXE has PTM site that is conserved
        """
        if 'Sequence Similarity Score' not in self.exploded_mxes.columns:
            sequence_similarity = []
            change_list = []
            for i, row in self.exploded_mxes.iterrows():
                if row['Flanking Sequence'] == row['Flanking Sequence']:
                    alt_sequence = row["Flanking Sequence"]
                    can_sequence = self.data.mapper.ptm_info.loc[row['Source of PTM'], 'Flanking Sequence']
                    score = int(pairwise2.align.globalxs(can_sequence.strip(' '), alt_sequence.strip(' '), -10, -2, score_only = True))
                    max_score = int(pairwise2.align.globalxs(can_sequence.strip(' '), can_sequence.strip(' '), -10, -2, score_only = True))
                    change = []
                    for i,e in enumerate(alt_sequence):
                        if len(can_sequence) == len(alt_sequence):
                            if can_sequence[i] != alt_sequence[i]:
                                change.append(i+1)
                    change_list.append(change)
                    sequence_similarity.append(score/max_score)
                else:
                    change_list.append(np.nan)
                    sequence_similarity.append(np.nan)

            self.exploded_mxes["Sequence Similarity Score"] = sequence_similarity
            self.exploded_mxes["Position Change"] = change_list

        conserved = self.exploded_mxes[self.exploded_mxes["Mapping Result"] == 'Success'].copy()
        mod_sizes = conserved.groupby('Modification').size().sort_values(ascending = False)
        mod_sizes = mod_sizes[mod_sizes >= min_instances]

        if gs is None:
            fig, ax = plt.subplots(figsize=(3, 6), nrows = mod_sizes.shape[0], sharex = True, sharey = True)
            fig.subplots_adjust(hspace=1)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(mod_sizes.shape[0], 1, subplot_spec=gs, hspace = 1)
            ax = [fig.add_subplot(gs2[i]) for i in range(mod_sizes.shape[0])]

        colors = sns.color_palette('colorblind', n_colors = mod_sizes.shape[0])
        for n,e in enumerate(mod_sizes.index):
            d1 = conserved[conserved["Modification"] == e]
            sample = conserved[conserved["Modification"] == e].shape[0]
            d1 = (d1["Position Change"].explode()-11).value_counts().sort_index(ascending = True)/sample
            if len(d1) == 0:
                ax[n].set_title(e + " (n = {})".format(sample), fontsize = self.data.axes_label_size)
            else:
                ax[n].bar(d1.index, d1.values, color = colors[n])
                ax[n].set_title(e + " (n = {})".format(sample), fontsize = self.data.axes_label_size)
            #ax[n].set_ylim([0,1])
            ax[n].set_xlim([-7.5,7.5])
            ax[n].set_xticks([-5,0,5])
            ax[n].set_xticklabels([-5,0,5])

        ax[int(n/2)].set_ylabel('Fraction of MXEs with Altered Residue at Corresponding Position', fontsize = self.data.axes_label_size)
        ax[mod_sizes.shape[0]-1].set_xlabel('Position in Flanking Sequence', fontsize = self.data.axes_label_size)

        if panel_save_dir is not None:
            plt.savefig(panel_save_dir + 'SFig10C_MXE_Positions.svg', bbox_inches = 'tight')

    def generate_figure(self, save_fig =True, save_panels = False, fig_type = 'svg'):
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, 0.8], wspace = 0.3)
        gs_col1 = gridspec.GridSpecFromSubplotSpec(3, 1, hspace = 0.2,height_ratios = [1,1,1], subplot_spec=gs[0])

        fig = plt.figure(figsize = (6.5, 7))

        #panel A
        ax00 = fig.add_subplot(gs_col1[0])
        self.PanelA(ax = ax00)
        fig.text(0.05, 0.95, 'A', fontsize=24, fontweight='bold', va='top', ha='left')

        #panel B
        ax10 = fig.add_subplot(gs_col1[1])
        self.PanelB(ax = ax10)
        fig.text(0.05, 0.7, 'B', fontsize=24, fontweight='bold', va='top', ha='left')

        #panel C
        self.PanelC(gs = gs[1], fig = fig)
        fig.text(0.5, 0.95, 'C', fontsize=24, fontweight='bold', va='top', ha='left')

        if save_fig:
            plt.savefig(self.data.fig_save_dir + f'SupplementaryFigure10.{fig_type}', bbox_inches = 'tight')

        if save_panels:
            self.PanelB(panel_save_dir = self.data.panel_save_dir)
            self.PanelC(panel_save_dir = self.data.panel_save_dir)





class SupplementaryFigure11:
    """
    Expansion on global altered flanking sequence analysis, which includes altered flanking rates compared to null model and the positions that are altered
    """
    def __init__(self, data, enriched_symbol = '*', depleted_symbol = '#'):
        #point to figure data class
        self.data = data

        self.mock_directory = self.data.analysis_dir + 'Null_Model/'

        #extract distance to borders for both modifications and residues
        ptm_distance = self.data.mapper.ptm_info.copy()
        ptm_distance['Distance to Closest Boundary (NC)'] = ptm_distance['Distance to Closest Boundary (NC)'].apply(lambda x: x.split(';'))
        ptm_distance = ptm_distance.explode('Distance to Closest Boundary (NC)').reset_index()
        ptm_distance['Distance to Closest Boundary (NC)'] = pd.to_numeric(ptm_distance['Distance to Closest Boundary (NC)'], errors = 'coerce')
        ptm_distance = ptm_distance.dropna(subset = 'Distance to Closest Boundary (NC)')
        ptm_distance = ptm_distance.drop_duplicates(subset = ['index', 'Distance to Closest Boundary (NC)'])
        self.ptm_distance = ptm_distance


        #save parameters
        self.enriched_symbol = enriched_symbol
        self.depleted_symbol = depleted_symbol

        #set other variables to none
        altered_flank_positions = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/PositionOfAlteredFlanks.csv')
        num_with_one_change = altered_flank_positions[altered_flank_positions['Altered_Positions'].apply(lambda x: '1.' in x)].shape[0]
        percent_with_one_change = num_with_one_change/altered_flank_positions.shape[0]
        print(f'{percent_with_one_change*100}% of altered flanking sequences have a change in the +1/-1 position, a total of {num_with_one_change} PTMs')
        self.altered_flank_positions = altered_flank_positions
        
        self.position_breakdown = pd.read_csv(self.data.analysis_dir + '/FlankingSequences/PositionBreakdown.csv', index_col = 0)


    def PanelA(self, ax = None, mod_list = ['Phosphorylation','Ubiquitination', 'Acetylation', 'Glycosylation', 'Methylation','Sumoylation', 'Dimethylation', 'Succinylation', 'Nitrosylation', 'Hydroxylation','Trimethylation', 'Crotonylation', 'Carboxylation', 'Palmitoylation', 'Sulfation'], legend = True, fig_save_dir = None):
        """
        Load null altered flanking rates and compare to the true altered flanking rates for each modification class

        Parameters
        ----------
        mod_list : list
            List of modifications to include in the analysis
        legend : bool
            Whether or not to include a legend in the plot
        save_figure : bool
            Whether or not to save the figure
        save_data : bool
            Whether or not to save the source data

        Returns
        -------
        None
        """
        #isolate ptms in isoforms and separate rows with multiple modification classes (separated by ;)
        exploded_isoform_ptms = self.data.mapper.isoform_ptms[self.data.mapper.isoform_ptms['Mapping Result'] == 'Success']
        exploded_isoform_ptms = exploded_isoform_ptms[exploded_isoform_ptms["Isoform Type"] == 'Alternative']
        exploded_isoform_ptms = exploded_isoform_ptms.drop_duplicates()

        #separate into mod specific rows
        exploded_isoform_ptms['Modification Class'] = exploded_isoform_ptms['Modification Class'].apply(lambda x: x.split(';'))
        exploded_isoform_ptms = exploded_isoform_ptms.explode('Modification Class').reset_index()


        #load altered flanking sequence rates in the null model, and combine with altered flanking rates for real ptms
        rate = []
        mod_results = []
        result_type = []
        for mod in mod_list:
            # open file with null altered flank rates
            with open(self.mock_directory + f'Null_Flank_Rates/{mod}.txt', 'r') as f:

                # write elements of list
                for line in f:
                    rate.append(float(line))
                    mod_results.append(mod)
                    result_type.append('Random')


            # close the file
            f.close()
            
            #grab real rates
            tmp = exploded_isoform_ptms[exploded_isoform_ptms['Modification Class'] == mod]
            rate.append(tmp[tmp['Conserved Flank (Size = 5)'] == 0].shape[0]/tmp.dropna(subset = 'Conserved Flank (Size = 5)').shape[0])
            mod_results.append(mod)
            result_type.append('Modified')
        
        #construct dataframe with null and real rates
        plt_data = pd.DataFrame({'Modification Class':mod_results, 'Residue Type': result_type, 'Constitutive Rate': rate})



        #perform "statistical" test: see fraction of null rates that are less than or greater than real rate, see if it is less than 0.05   
        mods = plt_data['Modification Class'].unique()
        comp_result = []
        for mod in mods:
            #grab real and random data for modification class of interest
            mod_data = plt_data[plt_data['Modification Class'] == mod]
            rand_data = mod_data[mod_data['Residue Type'] == 'Random']
            mod_rate = mod_data.loc[mod_data['Residue Type'] == 'Modified', 'Constitutive Rate'].values[0]
            #calculate the fraction of null rates that are less than or greater than real rate
            fraction_less = ((mod_rate >= rand_data['Constitutive Rate'])*1).sum()/rand_data.shape[0]
            fraction_more = ((mod_rate <= rand_data['Constitutive Rate'])*1).sum()/rand_data.shape[0]
            #determine if enrichment or depletion
            if fraction_less <= 0.05:
                comp_result.append('Depleted')
            elif fraction_more <= 0.05:
                comp_result.append('Enriched')
            else:
                comp_result.append('None')

        if ax is None:
            fig, ax = plt.subplots(figsize = (6.5,2.5))

        #blot barplot with null and real rates
        sns.barplot(x = 'Modification Class', y = 'Constitutive Rate', hue = 'Residue Type', data = plt_data.replace('Random', 'Null'), errorbar = 'sd', ax = ax) 
        #remove legend produced by seaborn from plot
        if not legend:
            ax.get_legend().remove()
        ticks = plt.xticks(rotation = 35, ha = 'right', fontsize = self.data.axes_label_size)
        ticks = plt.yticks(fontsize = self.data.axes_label_size-1)


        #add stat annotations
        for i in range(len(comp_result)):
            comp = comp_result[i]
            if comp == 'Depleted':
                stat = self.enriched_symbol
            elif comp == 'Enriched':
                stat = self.depleted_symbol
            else:
                stat = ''
                
            ax.annotate(stat, (i,0.055), ha = 'center', va = 'center', fontsize = 18, color = 'black')
        
        #format plot
        ax.set_ylim([0,0.06])
        ax.set_xlabel('')
        ax.set_ylabel('Fraction Altered', fontsize =self.data.axes_label_size)
        ax.legend(loc = (0.4,1.01), ncols = 2)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'SFig11A_NullFlank.svg', bbox_inches = 'tight', dpi = 300)
        return plt_data['Constitutive Rate'].max()



    def PanelB(self, ax = None, fig_save_dir = None):
        """
        Get fraction of altered flanks using different window sizes
        """
        self.conservation_scores = {}
        self.num_conserved ={}
        isoform_ptms = self.data.mapper.isoform_ptms[self.data.mapper.isoform_ptms["Isoform Type"] == 'Alternative']
        isoform_ptms = isoform_ptms.drop_duplicates()
        for flank_size in range(1,11):
            conserved_flanks = isoform_ptms[f'Conserved Flank (Size = {flank_size})'].dropna()
            #calculate fraction of ptms with conserved flank
            self.conservation_scores[flank_size] = sum(conserved_flanks)/len(conserved_flanks)
            self.num_conserved[flank_size] = len(conserved_flanks) - sum(conserved_flanks)

        if ax is None:
            fig, ax = plt.subplots(figsize =(1.7,2.3))

        plt_data = pd.Series(self.conservation_scores)
        ax.scatter(plt_data.index, 1-plt_data.values, c= 'black')
        ax.set_ylim([0,0.05])
        ax.set_xlabel('Flanking Sequence Window', fontsize = self.data.axes_label_size-1)
        ax.set_ylabel('Altered Regulatory\nRegion Rate', fontsize = self.data.axes_label_size-1)

        #add second y-axis which indicates the number of ptms at each flank size
        ax2 = ax.twinx()
        plt_data = pd.Series(self.num_conserved)
        ax2.scatter(plt_data.index, plt_data.values, c = 'black')
        ax2.set_ylim([0, 0.05*len(isoform_ptms.dropna(subset = 'Flanking Sequence'))])
        ax2.set_ylabel('Number of PTMs with Altered\nRegulatory Region', fontsize = self.data.axes_label_size-1)
        ax.tick_params(labelsize = 8)
        ax.tick_params(labelsize = 8)

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'SFig11B_Window.svg', bbox_inches = 'tight', dpi = 300)


    def PanelC(self, gs = None, fig = None, fig_save_dir = None):
        """
        Location that shows the location of the altered residues in the flanking sequence across all examples (top plot indicates whether it occurs on n- or c-terminal, bottom plot specifies the exact position)


        """
        if gs is None:
            fig, ax = plt.subplots(nrows = 2, figsize = (2.4,2.6), height_ratios = [0.6,1])
            fig.subplots_adjust(hspace = 1)
        else:
            gs2 = gridspec.GridSpecFromSubplotSpec(2, 1, subplot_spec=gs, hspace = 1)
            ax = [fig.add_subplot(gs2[i]) for i in range(2)]

        plt_data = self.altered_flank_positions.groupby('Location of Altered Flank').size()

    
        plt_data = plt_data[['N-term only', 'C-term only', 'Both']]
        ax[0].bar(plt_data.index, plt_data.values, color = 'gray')
        ax[0].set_xlabel('Location of Altered Region', fontsize = self.data.axes_label_size)
        ax[0].set_xticklabels(['N-term\nonly', 'C-term\nonly', 'Both'])
        ax[0].set_ylabel('# of PTMs', fontsize = self.data.axes_label_size)
        ax[0].set_yticks([0,5000,10000])
        ax[0].tick_params(labelsize = 8)


        plt_data = self.position_breakdown.copy()
        ax[1].bar(plt_data.index, plt_data['Number of PTMs'], color = 'gray')
        ax[1].set_xlim([-7.5,7.5])
        ax[1].set_xlabel('Position Relative to PTM', fontsize = self.data.axes_label_size)
        ax[1].set_ylabel('# of Changed\nResidues', fontsize = self.data.axes_label_size)
        ticks = ax[1].set_xticks([-7,-5,-3,-1,1,3,5,7])
        ax[1].tick_params(labelsize = 8)
        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + 'SFig11C_Position.svg', bbox_inches = 'tight', dpi = 300)

    def generate_figure(self, fig_save_dir = None, fig_type = 'svg'):
        gs = gridspec.GridSpec(2, 1, height_ratios = [1,1], hspace = 0.5)
        gs_row2 = gridspec.GridSpecFromSubplotSpec(1, 2, width_ratios = [1,2.5], wspace = 1.2, subplot_spec=gs[1])

        fig = plt.figure(figsize = (6.5, 7))


        #Panel A
        ax00 = fig.add_subplot(gs[0])
        self.PanelA(ax = ax00)
        fig.text(-0.02, 0.9, 'A', fontsize = 24, fontweight = 'bold')

        #Panel B
        ax01 = fig.add_subplot(gs_row2[0])
        self.PanelB(ax = ax01)
        fig.text(-0.02, 0.4, 'B', fontsize = 24, fontweight = 'bold')

        #Panel C
        self.PanelC(gs = gs_row2[1], fig = fig)
        fig.text(0.38, 0.4, 'C', fontsize = 24, fontweight = 'bold')

        if fig_save_dir is not None:
            plt.savefig(fig_save_dir + f'SupplementaryFigure11.{fig_type}', bbox_inches = 'tight', dpi = 300)


class SupplementaryTable1():
    """
    Generate table of non-constitutive PTMs and their functions
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data

    def create_table(self, save_dir):
        #grab function enrichment data, remove extra data, and rename columns to be clearer
        function = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Function_Enrichment.csv')
        function = function.drop(['Unnamed: 0', 'Subset'], axis = 1)
        function = function[['Function', 'Number Across All PTMs', 'Number Across PTM subset', 'Fraction in PTM subset', 'Odds Ratio', 'BH adjusted p']]
        function = function.rename({'Number Across All PTMs': 'Number of PTMs in Proteome with Function', 'Number Across PTM subset': "Number that are Constitutive", 'Fraction in PTM subset': 'Fraction that are Constitutive', 'BH adjusted p': 'Adjusted p-value'}, axis = 1)
        #repeat for process enrichment
        process = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Process_Enrichment.csv')
        process = process.drop(['Unnamed: 0', 'Subset'], axis = 1)
        process = process[['Process', 'Number Across All PTMs', 'Number Across PTM subset', 'Fraction in PTM subset', 'Odds Ratio', 'BH adjusted p']]
        process = process.rename({'Number Across All PTMs': 'Number of PTMs in Proteome associated with Process', 'Number Across PTM subset': "Number that are Constitutive", 'Fraction in PTM subset': 'Fraction that are Constitutive', 'BH adjusted p': 'Adjusted p-value'}, axis = 1)
        #repeat for kinase enrichment
        kinase = pd.read_csv(self.data.analysis_dir + '/Constitutive_Rates/Enrichment/Kinase_Enrichment.csv')
        kinase = kinase.drop(['Unnamed: 0', 'Subset'], axis = 1)
        kinase = kinase[['Kinase', 'Number Across All PTMs', 'Number Across PTM subset', 'Fraction in PTM subset', 'Odds Ratio', 'BH adjusted p']]
        kinase = kinase.rename({'Number Across All PTMs': 'Number of PTMs in Proteome associated with Process', 'Number Across PTM subset': "Number that are Constitutive", 'Fraction in PTM subset': 'Fraction that are Constitutive', 'BH adjusted p': 'Adjusted p-value'}, axis = 1)
        #save each table to a different sheet of the same excel file
        writer = pd.ExcelWriter(save_dir + 'SupplementaryTable1.xlsx', engine='openpyxl')
        function.to_excel(writer, sheet_name='Molecular Function Enrichment', index = False)
        process.to_excel(writer, sheet_name='Biological Process Enrichment', index = False)
        kinase.to_excel(writer, sheet_name='Kinase-Substrate Enrichment', index = False)

        writer.close()

    def calculateEnrichment(self, collapse_annotations = False, include_unknown = True, fishers = True):
        """
        Calculate enrichment of PTMs in constitutive or non-constititive ptms found in mapper.ptm_info, for each annotation and each modification class (including all PTMs)

        Parameters
        ----------
        mapper : PTM_mapper.PTMmapper object
            PTMmapper object that contains information about PTMs in proteome
        mod_groups : pandas.DataFrame
            Conversion table with information about modification types and their combined modification classes
        collapse_annotations : bool, optional
            If True, collapse PSP annotations into single entry for similar annotations. For example, 'cell motility, induced' and 'cell motility, inhibited' would be collapsed into 'cell motility'. The default is False.
        annotation_fname : str, optional
            Name of file that contains annotations from PhosphoSitePlus. The default is './Data/Regulatory_sites.gz'.

        Returns
        -------
        results : pandas.DataFrame
            Table with enrichment results for each annotation and each modification class (including all PTMs)

        """
        print('Getting enrichment of function for all ptms')
        #get list of constitutive and non-constitutive ptms
        constitutive_ptms = self.data.mapper.ptm_info[self.data.mapper.ptm_info['PTM Conservation Score'] == 1]
        non_constitutive_ptms = self.data.mapper.ptm_info[self.data.mapper.ptm_info['PTM Conservation Score'] != 1]
        

        function_table = stat_utils.constructPivotTable(self.data.mapper.ptm_info.reset_index(), self.data.regulatory, reference_col='ON_FUNCTION', collapse_on_similar = collapse_annotations, include_unknown=include_unknown)
        process_table = stat_utils.constructPivotTable(self.data.mapper.ptm_info.reset_index(),self.data.regulatory, reference_col='ON_PROCESS', collapse_on_similar = collapse_annotations, include_unknown=include_unknown)
        
        #enrichment, overall
        if fishers:
            self.function_enrichment = stat_utils.generate_site_enrichment(constitutive_ptms.index.values, function_table, subset_name = 'Constitutive', type = 'Function', fishers = fishers)
            self.process_enrichment = stat_utils.generate_site_enrichment(constitutive_ptms.index.values, process_table, subset_name = 'Constitutive', type = 'Process', fishers = fishers)
        else:
            cons_function_enrichment = stat_utils.generate_site_enrichment(constitutive_ptms.index.values, function_table, subset_name = 'Constitutive', type = 'Function', fishers = fishers)
            noncons_function_enrichment = stat_utils.generate_site_enrichment(non_constitutive_ptms.index.values, function_table, subset_name = 'Non-Constitutive', type = 'Function', fishers = fishers)
            self.function_enrichment = pd.concat([cons_function_enrichment, noncons_function_enrichment])   

            cons_process_enrichment = stat_utils.generate_site_enrichment(constitutive_ptms.index.values, process_table, subset_name = 'Constitutive', type = 'Process', fishers = fishers)
            noncons_process_enrichment = stat_utils.generate_site_enrichment(non_constitutive_ptms.index.values, process_table, subset_name = 'Non-Constitutive', type = 'Process', fishers = fishers)
            self.process_enrichment = pd.concat([cons_process_enrichment, noncons_process_enrichment])

    def get_nonconstitutive_table(self):
        """
        Process ptm_info and alternative ptm data to extract non-constitutive ptms and the transcripts/exons in which they are located
        """
        #extract non-constitutive ptms from ptm info 
        processed_ptms = self.data.mapper.ptm_info.copy()
        processed_ptms = processed_ptms[processed_ptms['PTM Conservation Score'] != 1]
        #add gene name to ptm info
        translator = config.translator[['Gene name', 'UniProtKB/Swiss-Prot ID']].dropna(subset = 'UniProtKB/Swiss-Prot ID').drop_duplicates()
        processed_ptms = processed_ptms.merge(translator, left_on = 'Protein', right_on = 'UniProtKB/Swiss-Prot ID', how = 'left')
        #combine residue and ptm location column
        processed_ptms['PTM Location (AA)'] = processed_ptms['Residue'] + processed_ptms['PTM Location (AA)'].astype(str)
        #extract columns of interest and rename
        processed_ptms = processed_ptms[['Gene name', 'PTM Location (AA)', 'Modification', 'Transcripts', 'Exon stable ID', 'Conserved Transcripts', 'Lost Transcripts']]
        processed_ptms = processed_ptms.rename({'Gene name':'Gene', 'Transcripts':'Canonical Transcript(s)', 'Exon stable ID': 'Exons containing PTM', 'Conserved Transcripts':'Alternative Transcripts with PTM', 'Lost Transcripts': 'Alternative Transcripts Excluding PTM'}, axis = 1)


        #grab alternative exon/PTM informatino
        alt_ptms = self.data.mapper.alternative_ptms
        alt_ptms = alt_ptms[alt_ptms['Mapping Result'] == 'Success']
        alt_ptms = alt_ptms.dropna(subset = ['Exon ID (Alternative)'])
        alt_ptms = alt_ptms.groupby('Source of PTM')['Exon ID (Alternative)'].agg(';'.join)

        #add alternative exons with ptms to main data
        processed_ptms = processed_ptms.join(alt_ptms)
        processed_ptms['Exons containing PTM'] = processed_ptms.apply(lambda x: '; '.join(np.unique(x['Exons containing PTM'].split(';') + x['Exon ID (Alternative)'].split(';'))) if x['Exon ID (Alternative)'] == x['Exon ID (Alternative)'] else '; '.join(np.unique(x['Exons containing PTM'].split(';'))), axis = 1)
        processed_ptms = processed_ptms.drop('Exon ID (Alternative)', axis = 1)

        # replace isoforms with ptms with transcripts with ptms 
        processed_ptms['Alternative Transcripts with PTM'] = self.convert_isoform_to_transcript(processed_ptms['Alternative Transcripts with PTM'])
        processed_ptms['Alternative Transcripts Excluding PTM'] = self.convert_isoform_to_transcript(processed_ptms['Alternative Transcripts Excluding PTM'])

        return processed_ptms
    
    def get_enrichment_table(self, type = 'Function'):
        pass

class SupplementaryTable2():
    """
    Generate processed flanking sequence table which compares canonical and alternative sequences (differences colored)
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data
        #make dict to convert column number to excel column
        self.excel_column_converter = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E',5:'F',6:'G', 7:'H', 8:'I',9:'J',10:'K',11:'L',12:'M',13:'N',14:'O',15:'P',16:'Q', 17:'R'}


    def format_table(self, flank_size = 5, save_dir = None):

        #grab alternative data, restrict to those with altered flanks
        save_data = self.data.mapper.isoform_ptms[(self.data.mapper.isoform_ptms['Conserved Flank (Size = 5)'] == 0) & (self.data.mapper.isoform_ptms['Mapping Result'] == 'Success')].copy()
        save_data = save_data[save_data['Isoform Type'] == 'Alternative']
        #add transcripts associated with each isoform
        save_data['Protein'] = save_data['Source of PTM'].apply(lambda x: x.split('-')[0]) 
        save_data = save_data[['Source of PTM', 'Protein', 'Modification', 'Isoform ID', 'Flanking Sequence', 'Isoform Type']]
        #explode out source of ptm information
        save_data['Source of PTM'] = save_data['Source of PTM'].apply(lambda x: x.split(';'))
        save_data = save_data.explode('Source of PTM')
        

        #grab canonical info for same ptms
        ptms_with_altered_flank = save_data['Source of PTM'].unique()
        canonical_info = self.data.mapper.ptm_info.loc[ptms_with_altered_flank].reset_index()
        canonical_info = canonical_info[canonical_info['Isoform Type'] == 'Canonical']
        canonical_info = canonical_info.rename({'index': 'Source of PTM', 'Protein':'Isoform ID'}, axis = 1)
        #set isoform id to be <UniProtID>-1
        canonical_info['Protein'] = canonical_info['Isoform ID'].apply(lambda x: x.split('-')[0])
        canonical_info = canonical_info[['Source of PTM', 'Protein', 'Modification', 'Isoform ID', 'Flanking Sequence']]
        canonical_info['Isoform Type'] = 'Canonical'

        save_data = save_data[save_data['Source of PTM'].isin(canonical_info['Source of PTM'])]

        #remove ptm source rows from alternative if not associated with canonical
        save_data = pd.concat([save_data, canonical_info])

        #add transcripts associated with alternative isoform
        isoforms = self.data.mapper.isoforms[['Isoform ID', 'Transcript stable ID']].drop_duplicates(subset = 'Isoform ID')
        save_data = save_data.merge(isoforms, on = 'Isoform ID', how = 'left')
        save_data = save_data.rename({'Transcript stable ID': 'Transcripts associated with Isoform'}, axis = 1)

        #add gene names to save data
        gene_names = config.translator[['Gene name', 'UniProtKB/Swiss-Prot ID']].drop_duplicates().dropna()
        save_data = save_data.merge(gene_names, left_on = 'Protein', right_on = 'UniProtKB/Swiss-Prot ID', how = 'left')

        #reorder columns
        save_data = save_data.drop('Protein', axis = 1)
        save_data= save_data[['Gene name', 'UniProtKB/Swiss-Prot ID', 'Source of PTM', 'Modification', 'Isoform ID', 'Transcripts associated with Isoform', 'Isoform Type', 'Flanking Sequence']]
        
        #rare cases where there are multiple gene names aggregate
        save_data = save_data.groupby([col for col in save_data.columns if col != 'Gene name'], as_index = False).agg(';'.join)

        #sort to group by ptm and start with canonical isoform
        save_data = save_data.sort_values(['Gene name','Source of PTM', 'Isoform Type'], ascending = [True, False, False])


        #isolate each flanking sequence position into its own column
        save_data['Flanking Sequence'] = save_data['Flanking Sequence'].apply(lambda x: list(x))
        #unpack lists into columns
        new_data = []
        for item in save_data['Flanking Sequence'].values:
            new_data.append(item)
        new_data = pd.DataFrame(new_data, index = save_data.index, columns = ['-10','-9', '-8','-7','-6', '-5', '-4', '-3', '-2', '-1','0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10'])
        #add split columns back to save_data
        save_data = pd.concat([save_data, new_data], axis = 1)
        save_data = save_data.drop('Flanking Sequence', axis = 1)

        #specifiy flank size and remove columns that are outside that window (max 10)
        cols_to_drop = [col for col in new_data.columns if abs(int(col)) > flank_size ]
        save_data = save_data.drop(cols_to_drop, axis = 1)

        #save dataframe prior to multindexing
        self.processed_data = save_data.copy()

        #create multiindex for future merging
        cols_to_index = [col for col in save_data.columns if col not in new_data.columns]
        save_data.index = pd.MultiIndex.from_frame(save_data[cols_to_index])
        save_data = save_data.drop(cols_to_index, axis = 1)

        #save to excel file
        if save_dir is not None:
            save_data.to_excel(save_dir + 'SupplementaryTable2.xlsx')

        self.save_data = save_data

    def find_cells_with_changed_data(self):
        #initialize list to store cell data
        cells_with_changed_data = []
        row_number = 2 #first row with data in excel file
        for ptm_data in self.processed_data['Source of PTM'].unique():
            #separate canonical and alternative information
            ptm_data = self.processed_data[self.processed_data['Source of PTM'] == ptm_data]
            canonical_row = ptm_data[ptm_data['Isoform Type'] == 'Canonical'].squeeze()
            alternative_rows = ptm_data[ptm_data['Isoform Type'] == 'Alternative']
            row_number += 1  #add one to row number since moving to next row
            for i, row in alternative_rows.iterrows():
                for i in range(len(row)):
                    if row.index[i] not in ['Gene name', 'UniProtKB/Swiss-Prot ID', 'Source of PTM','Modification', 'Isoform ID', "Isoform Type", 'Transcripts associated with Isoform'] and row[i] != canonical_row[i]:
                        cells_with_changed_data.append(self.excel_column_converter[i] + str(row_number))
                row_number += 1 #add one to row number since moving to next row

        return cells_with_changed_data

    def color_sheet(self, save_dir = None):
        if save_dir is None:
            save_dir = self.data.table_save_dir

        cells_with_changed_data = self.find_cells_with_changed_data()

        book = load_workbook(save_dir + 'SupplementaryTable2.xlsx')
        sheet = book.active
        for cell in cells_with_changed_data:
            sheet[cell].fill = PatternFill(start_color=	"FF7F7F", end_color="FF7F7F", fill_type = "solid")

        book.save(save_dir + 'SupplementaryTable2.xlsx')

    def generate_table(self, flank_size = 5, save_dir = None):

        if save_dir is None:
            raise ValueError('Must specify save directory')

        self.format_table(flank_size = flank_size, save_dir = save_dir)

        #save to excel file
        self.save_data.to_excel(save_dir + 'SupplementaryTable2.xlsx')

        #format excel file to color specific cells
        self.color_sheet(save_dir = save_dir)

class SupplementaryTable3:
    """
    Tables of PTMs regulated by ESRP1
    """
    def __init__(self, data):
        #point to figure data class
        self.data = data

        if self.data.prostate is None:
            self.data.prostate = constructCancerDict(self.data, 'PRAD', cutoff = 1, alpha = 0.05, min_effect_size=0.25)
        
        if self.data.breast is None:
            self.data.breast = constructCancerDict(self.data, 'Breast', cutoff = 1, alpha = 0.05, min_effect_size=0.25)
        
        if self.data.renal is None:
            self.data.renal = constructCancerDict(self.data, 'Renal', cutoff = 1, alpha = 0.05, min_effect_size=0.25)

        self.prostate_table = self.process_data('PRAD')
        self.breast_table = self.process_data('Breast')
        self.renal_table = self.process_data('Renal')

    def process_data(self, tissue = 'PRAD'):
        if tissue == 'PRAD':
            tissue_data = self.data.prostate
        elif tissue == 'Breast':
            tissue_data = self.data.breast
        elif tissue == 'Renal':
            tissue_data = self.data.renal
        else:
            raise ValueError('Invalid tissue type, currently only supports PRAD, Breast, and Renal')

        #extract ptm table, residue and modification information
        stable3 = tissue_data['Significant PTMs'].copy()
        stable3['Residue'] = stable3['PTM'].str.split('_').str[1]
        stable3 = stable3.merge(self.data.mapper.ptm_info['Modification'].reset_index(), on = 'PTM')

        #grab high and low patients and calculate median psi for each group
        high_patients = tissue_data['ESRP1-high patients']
        low_patients = tissue_data['ESRP1-low patients']
        stable3['Median PSI (ESRP1 High)'] = stable3[[pat for pat in high_patients if pat in stable3.columns]].median(axis = 1)
        stable3['Median PSI (ESRP1 Low)'] = stable3[[pat for pat in low_patients if pat in stable3.columns]].median(axis = 1)

        #extract only certain columns and rename
        stable3 = stable3[['symbol','exons','Residue','Modification', 'ESRP1','Median PSI (ESRP1 Low)', 'Median PSI (ESRP1 High)', 'Adj p', 'Effect Size']]
        stable3 = stable3.rename({'symbol':'Gene', 'exons':'Exon', 'ESRP1':'Group with Higher PSI', 'Adj p': 'Adjusted p-value'}, axis = 1)
        return stable3
    
    def create_low_table(self):
        """
        Extract prostate ptms that are included at a higher rate in the ESRP1 low group
        """
        low_table = self.prostate_table[self.prostate_table['Group with Higher PSI'] == 'Low'].copy()
        low_table = low_table.drop('Group with Higher PSI', axis = 1)
        return low_table
    
    def create_high_table(self):
        """
        Extract prostate ptms that are included at a higher rate in the ESRP1 high group
        """
        high_table = self.prostate_table[self.prostate_table['Group with Higher PSI'] == 'High'].copy()
        high_table = high_table.drop('Group with Higher PSI', axis = 1)
        return high_table
    
    def create_common_ptm_table(self):
        """
        Extract ptms that identified as ESRP1 regulated in prostate, breast, and renal
        """
        common_table = self.prostate_table.copy()
        #check if the same ptm is found for breast
        common_table = common_table[(common_table['Gene']+'_'+common_table['Residue']).isin((self.breast_table['Gene']+'_'+self.breast_table['Residue']))]
        #check if the same ptm is found for renal
        common_table = common_table[(common_table['Gene']+'_'+common_table['Residue']).isin((self.renal_table['Gene']+'_'+self.renal_table['Residue']))]
        common_table = common_table[['Gene', 'Exon', 'Residue', 'Modification']]
        return common_table
    
    def create_progressive_ptm_table(self):
        """
        Extract ptms that identified as ESRP1 regulated in prostate, breast, and renal
        """
        progressive_table = self.prostate_table.copy()
        #check if the same ptm is found for breast
        progressive_table = progressive_table[(progressive_table['Gene']+'_'+progressive_table['Residue']).isin((self.breast_table['Gene']+'_'+self.breast_table['Residue']))]
        #check if the same ptm is found for renal
        progressive_table = progressive_table[~(progressive_table['Gene']+'_'+progressive_table['Residue']).isin((self.renal_table['Gene']+'_'+self.renal_table['Residue']))]
        progressive_table = progressive_table[['Gene', 'Exon', 'Residue', 'Modification']]
        return progressive_table
    
    def create_worksheet(self, save_dir = None):
        if save_dir is None:
            save_dir = self.data.table_save_dir
        else:
            save_dir = save_dir

        low_table = self.create_low_table()
        high_table = self.create_high_table()
        common_table = self.create_common_ptm_table()
        progressive_table = self.create_progressive_ptm_table()

        #save each table to a different sheet of the same excel file
        writer = pd.ExcelWriter(save_dir + 'SupplementaryTable3.xlsx', engine='openpyxl')
        low_table.to_excel(writer, sheet_name='ESRP1-Low Included PTMs', index = False)
        high_table.to_excel(writer, sheet_name='ESRP1-High Included PTMs', index = False)
        common_table.to_excel(writer, sheet_name='Consistent Regulation by ESRP1', index = False)
        progressive_table.to_excel(writer, sheet_name='Oncogenic ESRP1 Regulation', index = False)

        writer.close()
     

def range_normalize(x, min_val, max_val, desired_range = [0,1]):
    """
    Normalize a value to a desired range

    Parameters
    ----------
    x : float
        value to normalize
    min_val : float
        minimum value of range to normalize to
    max_val : float
        maximum value of range to normalize to
    desired_range : list, optional
        desired range to normalize to. The default is [0,1].

    Returns
    -------
    float
        normalized value

    """
    return (x-min_val)/(max_val-min_val)*(desired_range[1]-desired_range[0])+desired_range[0]
        
def constructCancerDict(data, tissue = 'PRAD', cutoff = 1, alpha = 0.05, min_effect_size = 0.25, load_oncoprint = False):
    cancer_dict = {}
    #get regulated ptms
    ptms = pd.read_csv(data.analysis_dir + f'TCGA/{tissue}_ESRP1_PTMs.csv')
    cancer_dict['PTM'] = ptms

    #get signficant PTMs
    sig_ptms = ptms[(ptms['Adj p'] <= alpha) & (ptms['Effect Size'] >= min_effect_size)].copy()
    sig_ptms = sig_ptms.drop_duplicates(subset = ['PTM', 'ESRP1'])
    sig_ptms = sig_ptms.drop_duplicates(subset = 'PTM', keep = False) #remove PTMs that are significant in both directions
    cancer_dict['Significant PTMs'] = sig_ptms

    #get psi info
    psi = pd.read_csv(data.database_dir + f'TCGA/{tissue}/PSI_download_{tissue}.txt', sep = '\t')
    psi = psi[psi['splice_type'].isin(['ES','RI','AD','AA'])].copy()
    cancer_dict['PSI'] = psi

    #get ESRP1 expression, z-score normalized from CBio
    mrna_expression = pd.read_csv(data.database_dir + f"TCGA/{tissue}/{tissue}_ESRP1_expression.txt", sep='\t') 
    mrna_expression['SAMPLE_ID'] = mrna_expression['SAMPLE_ID'].str.replace('-','_')
    mrna_expression['SAMPLE_ID'] = mrna_expression['SAMPLE_ID'].str.strip('_01')
    mrna_expression.index = mrna_expression['SAMPLE_ID']
    mrna_expression = mrna_expression.drop(['SAMPLE_ID','STUDY_ID'], axis = 1)
    cancer_dict['ESRP1'] = mrna_expression

    #get which patients are high and low ESRP1
    cancer_dict['ESRP1-low patients'] = mrna_expression.loc[mrna_expression['ESRP1'] < -1*cutoff].index.values

    cancer_dict['ESRP1-high patients'] = mrna_expression.loc[mrna_expression['ESRP1'] > cutoff].index.values

    if load_oncoprint:
        cancer_dict['Oncoprint'] = pd.read_csv(data.database_dir + f'TCGA/{tissue}/PATIENT_DATA_oncoprint.tsv', sep = '\t')

    return cancer_dict





